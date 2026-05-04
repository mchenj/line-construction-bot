"""
Admin control center for LINE Construction Bot.

Routes:
- GET  /admin
- GET  /admin/api/overview
- GET  /admin/download/{kind}
- POST /admin/upload/{kind}
- POST /admin/trigger/weekly
- GET  /admin/check_fonts
"""

from __future__ import annotations

import os
import json
import subprocess
import warnings
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from html import escape
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse

try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover - Python always has zoneinfo on Railway's current stack.
    ZoneInfo = None

try:
    from supabase import create_client
except Exception:  # pragma: no cover - lets the admin page render even if deps are missing locally.
    create_client = None


_THIS_DIR = Path(__file__).parent
_DATA_DIR = _THIS_DIR / "data"
_DATA_DIR.mkdir(exist_ok=True)

DATA_FILES = {
    "plan": (_DATA_DIR / "construction_plan.xlsx", "Construction Plan"),
    "cm": (_DATA_DIR / "cm_personnel.xlsx", "CM Personnel"),
}

PROJECT_DISPLAY_NAME = "โครงการพัฒนาพื้นที่ชุมชนหัวรอและพื้นที่ต่อเนื่อง ตำบลหัวรอ อำเภอเมืองหัวรอฯ"
PLAN_BUDGET_SHEET = "การเบิกจ่ายงบประมาณ"
PLAN_PROGRESS_SHEET = "แผน - ผล ประจำสัปดาห์"

router = APIRouter(prefix="/admin", tags=["admin"])
public_router = APIRouter(tags=["public"])


def _now_bkk() -> datetime:
    if ZoneInfo:
        try:
            return datetime.now(ZoneInfo("Asia/Bangkok"))
        except Exception:
            pass
    return datetime.now(timezone(timedelta(hours=7)))


def _check_token(token: str):
    expected = os.getenv("ADMIN_TOKEN", "")
    if not expected:
        raise HTTPException(403, "ADMIN_TOKEN is not configured.")
    if token != expected:
        raise HTTPException(401, "Invalid admin token.")


def _file_info(path: Path) -> dict:
    if not path.exists():
        return {"exists": False, "size": 0, "modified": None, "age_hours": None}
    st = path.stat()
    modified = datetime.fromtimestamp(st.st_mtime, tz=_now_bkk().tzinfo)
    age_hours = max(0, (_now_bkk() - modified).total_seconds() / 3600)
    return {
        "exists": True,
        "size": st.st_size,
        "size_label": _format_bytes(st.st_size),
        "modified": modified.strftime("%Y-%m-%d %H:%M"),
        "age_hours": age_hours,
    }


def _format_bytes(size: int) -> str:
    value = float(size)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{value:.1f} GB"


def _parse_iso_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        text = str(value).replace("Z", "+00:00")
        parsed = datetime.fromisoformat(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(_now_bkk().tzinfo)
    except Exception:
        return None


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _short(text: object, length: int = 92) -> str:
    clean = " ".join(str(text or "").split())
    if len(clean) <= length:
        return clean
    return clean[: max(0, length - 1)].rstrip() + "..."


def _get_supabase_client():
    if create_client is None:
        return None, "The supabase package is not available."
    url = os.getenv("SUPABASE_URL", "").strip()
    key = os.getenv("SUPABASE_KEY", "").strip()
    if not url or not key:
        return None, "SUPABASE_URL or SUPABASE_KEY is missing."
    try:
        return create_client(url, key), None
    except Exception as exc:
        return None, str(exc)


def _safe_query(label: str, fn):
    try:
        result = fn()
        return list(result.data or []), None
    except Exception as exc:
        return [], f"{label}: {exc}"


def _env_status() -> dict:
    weekly_enabled = os.getenv("WEEKLY_CRON_ENABLED", "false").lower() in ("true", "1", "yes")
    railway_name = os.getenv("RAILWAY_SERVICE_NAME") or os.getenv("RAILWAY_PROJECT_NAME") or "Local runtime"
    railway_domain = os.getenv("RAILWAY_PUBLIC_DOMAIN") or os.getenv("RAILWAY_STATIC_URL") or ""
    weekly_hour = _safe_int(os.getenv("WEEKLY_CRON_HOUR", "17"), 17)
    weekly_minute = _safe_int(os.getenv("WEEKLY_CRON_MINUTE", "0"), 0)
    return {
        "line_secret": bool(os.getenv("LINE_CHANNEL_SECRET", "").strip()),
        "line_token": bool(os.getenv("LINE_CHANNEL_ACCESS_TOKEN", "").strip()),
        "supabase_url": bool(os.getenv("SUPABASE_URL", "").strip()),
        "supabase_key": bool(os.getenv("SUPABASE_KEY", "").strip()),
        "admin_token": bool(os.getenv("ADMIN_TOKEN", "").strip()),
        "weekly_enabled": weekly_enabled,
        "weekly_schedule": (
            f"{os.getenv('WEEKLY_CRON_DAY', 'fri').upper()} "
            f"{weekly_hour:02d}:"
            f"{weekly_minute:02d}"
        ),
        "weekly_format": os.getenv("WEEKLY_CRON_FORMAT", "zip").upper(),
        "weekly_targets": len([x for x in os.getenv("WEEKLY_CRON_USER_IDS", "").split(",") if x.strip()]),
        "railway_name": railway_name,
        "railway_domain": railway_domain,
        "port": os.getenv("PORT", "local"),
        "project_name": os.getenv("PROJECT_NAME", PROJECT_DISPLAY_NAME),
    }


def _safe_int(value: object, fallback: int) -> int:
    try:
        return int(str(value).strip().split()[0])
    except Exception:
        return fallback


def _collect_dashboard() -> dict:
    now = _now_bkk()
    today = now.date()
    start_14 = today - timedelta(days=13)
    start_30 = today - timedelta(days=29)
    env = _env_status()
    client, client_error = _get_supabase_client()

    data = {
        "generated_at": now.strftime("%Y-%m-%d %H:%M:%S"),
        "today": today.isoformat(),
        "env": env,
        "files": {kind: {"label": label, **_file_info(path)} for kind, (path, label) in DATA_FILES.items()},
        "db_ok": False,
        "db_error": client_error,
        "query_errors": [],
        "daily": [],
        "line_reports": [],
        "activities": [],
        "images": [],
        "issues": [],
        "issue_comments": [],
        "issues_error": None,
    }

    if client is not None:
        daily, err = _safe_query(
            "daily_reports",
            lambda: client.table("daily_reports")
            .select(
                "work_date,total_workers,engineers,foremen,skilled_workers,"
                "laborers,equipment,water_level,weather_morning,report_status,updated_at"
            )
            .gte("work_date", start_30.isoformat())
            .order("work_date")
            .execute(),
        )
        data["daily"] = daily
        if err:
            data["query_errors"].append(err)

        line_reports, err = _safe_query(
            "line_reports",
            lambda: client.table("line_reports")
            .select("id,timestamp,user_id,message_type,work_date,raw_text,image_url")
            .gte("timestamp", (now - timedelta(days=30)).astimezone(timezone.utc).isoformat())
            .order("timestamp", desc=True)
            .limit(300)
            .execute(),
        )
        data["line_reports"] = line_reports
        if err:
            data["query_errors"].append(err)

        activities, err = _safe_query(
            "report_activities",
            lambda: client.table("report_activities")
            .select("work_date,activity_type,description,seq_no")
            .gte("work_date", start_30.isoformat())
            .order("work_date", desc=True)
            .limit(500)
            .execute(),
        )
        data["activities"] = activities
        if err:
            data["query_errors"].append(err)

        images, err = _safe_query(
            "report_images",
            lambda: client.table("report_images")
            .select("work_date,image_url,caption,taken_at")
            .gte("work_date", start_30.isoformat())
            .order("taken_at", desc=True)
            .limit(24)
            .execute(),
        )
        data["images"] = images
        if err:
            data["query_errors"].append(err)

        issues, err = _safe_query(
            "project_issues",
            lambda: client.table("project_issues")
            .select(
                "id,created_at,updated_at,work_date,title,description,area,owner,"
                "due_date,status,impact,next_action,source_channel,source_id"
            )
            .order("created_at", desc=True)
            .limit(80)
            .execute(),
        )
        data["issues"] = issues
        if err:
            data["issues_error"] = err

        comments, err = _safe_query(
            "project_issue_comments",
            lambda: client.table("project_issue_comments")
            .select("id,issue_id,created_at,author,comment")
            .order("created_at", desc=True)
            .limit(120)
            .execute(),
        )
        data["issue_comments"] = comments
        if err and not data["issues_error"]:
            data["issues_error"] = err

        data["db_ok"] = any([daily, line_reports, activities, images]) or not data["query_errors"]
        if data["db_ok"]:
            data["db_error"] = None

    data.update(_summarize_dashboard(data, start_14, start_30))
    return data


def _summarize_dashboard(data: dict, start_14: date, start_30: date) -> dict:
    today = _parse_date(data["today"]) or _now_bkk().date()
    line_reports = data["line_reports"]
    daily = data["daily"]
    activities = data["activities"]
    images = data["images"]
    issues = data.get("issues") or []
    issue_comments = data.get("issue_comments") or []

    by_day = {start_14 + timedelta(days=i): {"reports": 0, "workers": 0, "images": 0} for i in range(14)}
    recent_by_type = Counter()
    unique_users = set()
    today_messages = 0
    week_messages = 0

    for row in line_reports:
        work_date = _parse_date(row.get("work_date")) or _parse_iso_dt(row.get("timestamp")) and _parse_iso_dt(row.get("timestamp")).date()
        msg_type = row.get("message_type") or "unknown"
        recent_by_type[msg_type] += 1
        if row.get("user_id"):
            unique_users.add(row.get("user_id"))
        if work_date == today:
            today_messages += 1
        if work_date and work_date >= today - timedelta(days=6):
            week_messages += 1
        if work_date in by_day:
            by_day[work_date]["reports"] += 1

    latest_daily = None
    worker_total_7d = 0
    water_points = []
    for row in daily:
        work_date = _parse_date(row.get("work_date"))
        if work_date in by_day:
            by_day[work_date]["workers"] = int(row.get("total_workers") or 0)
        if work_date and work_date >= today - timedelta(days=6):
            worker_total_7d += int(row.get("total_workers") or 0)
        if row.get("water_level") is not None and work_date:
            try:
                water_points.append({"date": work_date.isoformat(), "value": float(row.get("water_level"))})
            except Exception:
                pass
        if work_date and (latest_daily is None or work_date > _parse_date(latest_daily.get("work_date"))):
            latest_daily = row

    for img in images:
        work_date = _parse_date(img.get("work_date"))
        if work_date in by_day:
            by_day[work_date]["images"] += 1

    activity_counts = Counter((row.get("activity_type") or "general") for row in activities)
    activity_top = activity_counts.most_common(8)
    active_date = today
    if latest_daily:
        active_date = _parse_date(latest_daily.get("work_date")) or today
    today_activities = _activity_items_for_date(activities, active_date)
    problem_report = _problem_report(activities, line_reports, issues)
    financial_report = _financial_report()
    progress_report = _progress_report(daily, activities, images, latest_daily)
    field_context = _field_context(line_reports, latest_daily, today_activities, images)
    issue_board = _issue_board(issues, issue_comments, today)

    plan_ok = data["files"]["plan"]["exists"]
    cm_ok = data["files"]["cm"]["exists"]
    health_items = [
        data["env"]["line_secret"],
        data["env"]["line_token"],
        data["env"]["supabase_url"],
        data["env"]["supabase_key"],
        data["env"]["admin_token"],
        data["db_ok"],
        plan_ok,
        cm_ok,
    ]
    health_score = round((sum(1 for item in health_items if item) / len(health_items)) * 100)

    return {
        "health_score": health_score,
        "metrics": {
            "today_messages": today_messages,
            "week_messages": week_messages,
            "active_users": len(unique_users),
            "photos_30d": len(images),
            "worker_total_7d": worker_total_7d,
            "activity_types": len(activity_counts),
            "open_issues": issue_board["open_count"],
        },
        "latest_daily": latest_daily,
        "daily_series": [
            {
                "date": day.isoformat(),
                "label": day.strftime("%d %b"),
                "reports": values["reports"],
                "workers": values["workers"],
                "images": values["images"],
            }
            for day, values in by_day.items()
        ],
        "message_mix": dict(recent_by_type),
        "activity_top": activity_top,
        "water_points": water_points[-14:],
        "recent_events": _recent_events(line_reports[:12]),
        "field_context": field_context,
        "today_activities": today_activities,
        "financial_report": financial_report,
        "progress_report": progress_report,
        "problem_report": problem_report,
        "issue_board": issue_board,
        "recent_photos": _recent_photos(images),
    }


def _activity_items_for_date(rows: list[dict], target: date) -> list[dict]:
    items = []
    for row in sorted(rows, key=lambda x: (str(x.get("work_date") or ""), int(x.get("seq_no") or 0))):
        if _parse_date(row.get("work_date")) != target:
            continue
        items.append(
            {
                "type": str(row.get("activity_type") or "general").replace("_", " ").title(),
                "text": _short(row.get("description") or "Activity recorded", 92),
                "seq": int(row.get("seq_no") or len(items) + 1),
                "date": target.isoformat(),
            }
        )
    return items[:6]


def _equipment_items(raw: object) -> list[dict]:
    if not raw:
        return []
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            return []
    if isinstance(raw, dict):
        raw = [raw]
    if not isinstance(raw, list):
        return []
    items = []
    for item in raw:
        if isinstance(item, dict):
            items.append(item)
    return items


def _equipment_count(raw: object) -> int:
    total = 0
    for item in _equipment_items(raw):
        qty = _number_or_none(item.get("qty"))
        total += int(qty if qty is not None else 1)
    return total


def _field_context(line_reports: list[dict], latest_daily: dict | None, activities: list[dict], images: list[dict]) -> dict:
    latest_event = _recent_events(line_reports[:1])
    workers = int((latest_daily or {}).get("total_workers") or 0)
    engineers = int((latest_daily or {}).get("engineers") or 0)
    equipment_count = _equipment_count((latest_daily or {}).get("equipment"))
    water = (latest_daily or {}).get("water_level")
    return {
        "latest_text": latest_event[0]["text"] if latest_event else "Waiting for the next LINE field update.",
        "latest_time": latest_event[0]["time"] if latest_event else "No live event yet",
        "work_date": str((latest_daily or {}).get("work_date") or ""),
        "workers": workers,
        "engineers": engineers,
        "equipment_count": equipment_count,
        "weather": (latest_daily or {}).get("weather_morning") or "Not reported",
        "water": "" if water is None else f"{float(water):.2f} m",
        "activity_count": len(activities),
        "photo_count": len(images),
    }


def _open_issue_rows(issues: list[dict]) -> list[dict]:
    return [
        row for row in issues
        if str(row.get("status") or "open").lower() not in {"resolved", "closed"}
    ]


def _issue_board(issues: list[dict], comments: list[dict], today: date) -> dict:
    open_rows = _open_issue_rows(issues)
    by_status = Counter(str(row.get("status") or "open").lower() for row in issues)
    by_impact = Counter(str(row.get("impact") or "medium").lower() for row in open_rows)
    overdue = []
    due_soon = []
    for row in open_rows:
        due = _parse_date(row.get("due_date"))
        if not due:
            continue
        if due < today:
            overdue.append(row)
        elif due <= today + timedelta(days=3):
            due_soon.append(row)

    comments_by_issue: dict[int, list[dict]] = {}
    for comment in comments:
        issue_id = comment.get("issue_id")
        if issue_id is None:
            continue
        try:
            issue_id = int(issue_id)
        except Exception:
            continue
        comments_by_issue.setdefault(issue_id, []).append(comment)

    return {
        "issues": issues,
        "open_issues": open_rows,
        "comments_by_issue": comments_by_issue,
        "open_count": len(open_rows),
        "total_count": len(issues),
        "overdue_count": len(overdue),
        "due_soon_count": len(due_soon),
        "by_status": dict(by_status),
        "by_impact": dict(by_impact),
        "source_ready": not bool(issues is None),
    }


def _problem_report(activities: list[dict], line_reports: list[dict], issues: list[dict] | None = None) -> dict:
    open_issues = _open_issue_rows(issues or [])
    if open_issues:
        latest = open_issues[0]
        status = "Attention"
        if len(open_issues) <= 2:
            status = "Watch"
        return {
            "count": len(open_issues),
            "status": status,
            "latest": {
                "date": str(latest.get("work_date") or ""),
                "text": _short(latest.get("title") or latest.get("description") or "Open project issue.", 120),
            },
        }

    keywords = (
        "problem", "issue", "risk", "delay", "blocked", "struggle",
        "ปัญหา", "ล่าช้า", "ติด", "ขัดข้อง", "เสีย", "รอ", "แก้ไข",
    )
    candidates = []
    for row in list(activities) + list(line_reports):
        text = " ".join(str(row.get(k) or "") for k in ("description", "raw_text", "activity_type"))
        lowered = text.lower()
        if any(k.lower() in lowered for k in keywords):
            candidates.append(
                {
                    "date": str(row.get("work_date") or ""),
                    "text": _short(text, 120),
                }
            )
    severity = "Clear"
    if len(candidates) >= 3:
        severity = "Watch"
    if len(candidates) >= 6:
        severity = "Attention"
    return {
        "count": len(candidates),
        "status": severity,
        "latest": candidates[0] if candidates else {"date": "", "text": "No problem keywords detected in recent reports."},
    }


def _open_plan_workbook():
    path = DATA_FILES["plan"][0]
    if not path.exists():
        return None, f"{path.name} is missing."
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return None, f"openpyxl is not available: {exc}"
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            return load_workbook(path, data_only=True, read_only=True), None
    except Exception as exc:
        return None, f"Cannot read {path.name}: {exc}"


def _find_sheet(workbook, preferred: str):
    if preferred in workbook.sheetnames:
        return workbook[preferred]
    compact = preferred.replace(" ", "")
    for sheet in workbook.worksheets:
        if compact in sheet.title.replace(" ", ""):
            return sheet
    return None


def _financial_report() -> dict:
    workbook, error = _open_plan_workbook()
    if workbook is not None:
        try:
            sheet = _find_sheet(workbook, PLAN_BUDGET_SHEET)
            if sheet is None:
                error = f"Sheet {PLAN_BUDGET_SHEET} was not found."
            else:
                rows = []
                total_row = None
                for row_no in range(1, min(sheet.max_row, 30) + 1):
                    label = str(sheet.cell(row_no, 2).value or "").strip()
                    budget = _number_or_none(sheet.cell(row_no, 3).value)
                    spent = _number_or_none(sheet.cell(row_no, 5).value)
                    remaining = _number_or_none(sheet.cell(row_no, 7).value)
                    if not label or label in {"สรุปการใช้จ่าย", "ตามปีงบประมาณ"}:
                        continue
                    if budget is None and spent is None and remaining is None:
                        continue
                    item = {
                        "label": label,
                        "budget": budget or 0,
                        "spent": spent or 0,
                        "remaining": remaining or 0,
                    }
                    if label == "รวม" and total_row is None:
                        total_row = item
                    else:
                        rows.append(item)

                if total_row is None and rows:
                    total_row = {
                        "label": "Total",
                        "budget": sum(item["budget"] for item in rows),
                        "spent": sum(item["spent"] for item in rows),
                        "remaining": sum(item["remaining"] for item in rows),
                    }

                if total_row and total_row["budget"] > 0:
                    budget = float(total_row["budget"])
                    spent = float(total_row["spent"])
                    remaining = float(total_row["remaining"])
                    used = max(0, min(100, round((spent / budget) * 100, 1)))
                    return {
                        "connected": True,
                        "headline": f"{used:.1f}% disbursed",
                        "spent_label": _money(spent),
                        "budget_label": _money(budget),
                        "remaining_label": _money(remaining),
                        "committed_label": _money(remaining),
                        "percent": used,
                        "rows": rows,
                        "total": total_row,
                        "note": f"{DATA_FILES['plan'][0].name} / {PLAN_BUDGET_SHEET}",
                    }
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    budget = _safe_float(os.getenv("PROJECT_BUDGET"))
    spent = _safe_float(os.getenv("PROJECT_SPENT"))
    committed = _safe_float(os.getenv("PROJECT_COMMITTED"))
    if budget <= 0:
        return {
            "connected": False,
            "headline": "Financial source pending",
            "spent_label": "-",
            "budget_label": "-",
            "remaining_label": "-",
            "committed_label": "-",
            "percent": 0,
            "rows": [],
            "total": None,
            "note": error or "Set PROJECT_BUDGET, PROJECT_SPENT and PROJECT_COMMITTED or connect a finance table.",
        }
    used = max(0, min(100, round((spent / budget) * 100)))
    return {
        "connected": True,
        "headline": f"{used}% used",
        "spent_label": _money(spent),
        "budget_label": _money(budget),
        "remaining_label": _money(max(0, budget - spent)),
        "committed_label": _money(committed),
        "percent": used,
        "rows": [],
        "total": {"label": "Total", "budget": budget, "spent": spent, "remaining": max(0, budget - spent)},
        "note": "Budget values from Railway environment variables.",
    }


def _previous_project_period(reference: date) -> tuple[date, date]:
    if reference.day <= 7:
        prior_month_end = reference.replace(day=1) - timedelta(days=1)
        return prior_month_end.replace(day=24), prior_month_end
    if reference.day <= 15:
        return reference.replace(day=1), reference.replace(day=7)
    if reference.day <= 23:
        return reference.replace(day=8), reference.replace(day=15)
    return reference.replace(day=16), reference.replace(day=23)


def _normal_plan_date(value: object) -> date | None:
    parsed = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif value:
        parsed = _parse_date(str(value)[:10])
    if parsed and parsed.year >= 2400:
        try:
            parsed = parsed.replace(year=parsed.year - 543)
        except ValueError:
            return None
    return parsed


def _period_label(start: date, end: date) -> str:
    if start.month == end.month and start.year == end.year:
        return f"{start.strftime('%d')}-{end.strftime('%d %b %Y')}"
    return f"{start.strftime('%d %b')}-{end.strftime('%d %b %Y')}"


def _progress_report(daily: list[dict], activities: list[dict], images: list[dict], latest_daily: dict | None) -> dict:
    reported_days = len({str(row.get("work_date")) for row in daily if row.get("work_date")})
    period_start, period_end = _previous_project_period(_now_bkk().date())
    period = _period_label(period_start, period_end)
    workbook, error = _open_plan_workbook()
    if workbook is not None:
        try:
            sheet = _find_sheet(workbook, PLAN_PROGRESS_SHEET)
            if sheet is None:
                error = f"Sheet {PLAN_PROGRESS_SHEET} was not found."
            else:
                exact_row = None
                fallback_row = None
                fallback_end = None
                fallback_start = None
                for row_no in range(5, sheet.max_row + 1):
                    start = _normal_plan_date(sheet.cell(row_no, 4).value)
                    end = _normal_plan_date(sheet.cell(row_no, 5).value)
                    if not start or not end:
                        continue
                    if start == period_start and end == period_end:
                        exact_row = (row_no, start, end)
                        break
                    if start <= period_start and end >= period_end:
                        exact_row = (row_no, start, end)
                    if end <= period_end and (fallback_end is None or end > fallback_end):
                        fallback_row = row_no
                        fallback_start = start
                        fallback_end = end

                selected = exact_row
                if selected is None and fallback_row is not None:
                    selected = (fallback_row, fallback_start, fallback_end)

                if selected:
                    row_no, start, end = selected
                    plan = _number_or_none(sheet.cell(row_no, 8).value) or 0
                    actual = _number_or_none(sheet.cell(row_no, 10).value) or 0
                    variance = _number_or_none(sheet.cell(row_no, 12).value)
                    if variance is None:
                        variance = actual - plan
                    weekly_plan = _number_or_none(sheet.cell(row_no, 7).value) or 0
                    weekly_actual = _number_or_none(sheet.cell(row_no, 9).value) or 0
                    status = "On plan"
                    if variance > 0.05:
                        status = "Ahead"
                    elif variance < -0.05:
                        status = "Behind"
                    return {
                        "percent": max(0, min(100, round(actual, 1))),
                        "headline": f"{status} {abs(variance):.2f}%",
                        "status": status,
                        "period_label": _period_label(start, end),
                        "week_no": str(sheet.cell(row_no, 6).value or ""),
                        "plan_percent": round(plan, 2),
                        "actual_percent": round(actual, 2),
                        "variance_percent": round(variance, 2),
                        "weekly_plan_percent": round(weekly_plan, 2),
                        "weekly_actual_percent": round(weekly_actual, 2),
                        "reported_days": reported_days,
                        "activity_count": len(activities),
                        "photo_count": len(images),
                        "latest_status": str((latest_daily or {}).get("report_status") or "draft").title(),
                        "note": f"Previous project period from {DATA_FILES['plan'][0].name}: {period}.",
                    }
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    coverage = max(0, min(100, round((reported_days / 30) * 100)))
    latest_status = (latest_daily or {}).get("report_status") or "draft"
    return {
        "percent": coverage,
        "headline": "Progress source pending",
        "status": "Pending",
        "period_label": period,
        "week_no": "",
        "plan_percent": 0,
        "actual_percent": 0,
        "variance_percent": 0,
        "weekly_plan_percent": 0,
        "weekly_actual_percent": 0,
        "reported_days": reported_days,
        "activity_count": len(activities),
        "photo_count": len(images),
        "latest_status": str(latest_status).title(),
        "note": error or "Use this as field reporting coverage until plan progress is connected.",
    }


def _recent_photos(rows: list[dict]) -> list[dict]:
    photos = []
    for row in rows[:6]:
        photos.append(
            {
                "url": row.get("image_url"),
                "caption": _short(row.get("caption") or "Recent field photo", 72),
                "date": str(row.get("work_date") or ""),
            }
        )
    return photos


def _safe_float(value: object) -> float:
    try:
        text = str(value or "").replace(",", "").strip()
        return float(text) if text else 0.0
    except Exception:
        return 0.0


def _number_or_none(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text or text in {"-", "—"}:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _money(value: float) -> str:
    if abs(value) >= 1_000_000:
        return f"{value / 1_000_000:.1f}M"
    if abs(value) >= 1_000:
        return f"{value / 1_000:.1f}K"
    return f"{value:.0f}"


def _percent_label(value: object) -> str:
    number = _number_or_none(value)
    if number is None:
        return "-"
    return f"{number:.2f}%"


def _signed_percent_label(value: object) -> str:
    number = _number_or_none(value)
    if number is None:
        return "-"
    return f"{number:+.2f}%"


def _recent_events(rows: list[dict]) -> list[dict]:
    events = []
    for row in rows:
        dt = _parse_iso_dt(row.get("timestamp"))
        label = dt.strftime("%d %b %H:%M") if dt else str(row.get("timestamp") or "")
        message_type = str(row.get("message_type") or "message")
        events.append(
            {
                "time": label,
                "type": message_type.title(),
                "date": str(row.get("work_date") or ""),
                "text": _short(row.get("raw_text") or row.get("image_url") or "No text", 110),
            }
        )
    return events


def _status_word(ok: bool, paused: bool = False) -> str:
    if paused:
        return "Paused"
    return "Online" if ok else "Needs setup"


def _status_class(ok: bool, paused: bool = False) -> str:
    if paused:
        return "warn"
    return "ok" if ok else "bad"


def _render_metric(title: str, value: object, note: str, accent: str) -> str:
    return f"""
    <article class="metric" style="--accent:{escape(accent)}">
      <span class="metric-key">{escape(title)}</span>
      <strong>{escape(str(value))}</strong>
      <small>{escape(note)}</small>
    </article>
    """


def _render_report_chart(series: list[dict]) -> str:
    width, height = 720, 236
    left, top, bottom = 44, 22, 42
    plot_h = height - top - bottom
    plot_w = width - left - 20
    max_reports = max([int(item["reports"]) for item in series] + [1])
    max_workers = max([int(item["workers"]) for item in series] + [1])
    slot = plot_w / max(1, len(series))
    bar_w = max(14, min(28, slot * 0.45))
    parts = [
        f'<svg viewBox="0 0 {width} {height}" role="img" aria-label="Fourteen day report trend">'
        f'<rect width="{width}" height="{height}" fill="#fffaf0"/>'
    ]
    for i in range(5):
        y = top + plot_h * i / 4
        parts.append(f'<line x1="{left}" y1="{y:.1f}" x2="{width-12}" y2="{y:.1f}" stroke="#f1d7ad" stroke-width="1"/>')
    for idx, item in enumerate(series):
        x = left + idx * slot + (slot - bar_w) / 2
        report_h = (int(item["reports"]) / max_reports) * (plot_h - 8)
        worker_h = (int(item["workers"]) / max_workers) * (plot_h - 8)
        y1 = top + plot_h - report_h
        y2 = top + plot_h - worker_h
        parts.append(f'<rect x="{x:.1f}" y="{y1:.1f}" width="{bar_w:.1f}" height="{report_h:.1f}" rx="4" fill="#b91c1c"/>')
        parts.append(f'<rect x="{x + bar_w + 3:.1f}" y="{y2:.1f}" width="{max(5, bar_w * 0.32):.1f}" height="{worker_h:.1f}" rx="3" fill="#f59e0b"/>')
        label = escape(str(item["label"]))
        parts.append(
            f'<text x="{x + bar_w / 2:.1f}" y="{height-18}" text-anchor="middle" '
            f'font-size="11" fill="#7f1d1d">{label}</text>'
        )
    parts.append(f'<text x="{left}" y="15" font-size="12" fill="#7f1d1d">Reports and workers</text>')
    parts.append(f'<circle cx="{width-180}" cy="12" r="5" fill="#b91c1c"/><text x="{width-170}" y="16" font-size="11" fill="#7f1d1d">Messages</text>')
    parts.append(f'<circle cx="{width-95}" cy="12" r="5" fill="#f59e0b"/><text x="{width-85}" y="16" font-size="11" fill="#7f1d1d">Workers</text>')
    parts.append("</svg>")
    return "".join(parts)


def _render_activity_chart(activity_top: list[tuple[str, int]]) -> str:
    if not activity_top:
        activity_top = [("No activity yet", 1)]
    max_value = max(count for _, count in activity_top) or 1
    rows = []
    for name, count in activity_top:
        pct = max(5, round((count / max_value) * 100))
        rows.append(
            f"""
            <div class="bar-row">
              <span>{escape(str(name).replace("_", " ").title())}</span>
              <div class="bar-track"><i style="width:{pct}%"></i></div>
              <b>{count}</b>
            </div>
            """
        )
    return "".join(rows)


def _render_waterline(points: list[dict]) -> str:
    width, height = 560, 150
    if len(points) < 2:
        return (
            f'<svg viewBox="0 0 {width} {height}" role="img" aria-label="Water level chart">'
            f'<rect width="{width}" height="{height}" fill="#fffaf0"/>'
            f'<text x="24" y="78" font-size="13" fill="#7f1d1d">Waiting for water level data</text>'
            "</svg>"
        )
    values = [float(item["value"]) for item in points]
    min_v, max_v = min(values), max(values)
    span = max(max_v - min_v, 1)
    coords = []
    for idx, item in enumerate(points):
        x = 22 + idx * ((width - 44) / max(1, len(points) - 1))
        y = 20 + (max_v - float(item["value"])) / span * (height - 52)
        coords.append((x, y))
    poly = " ".join(f"{x:.1f},{y:.1f}" for x, y in coords)
    area = f"22,{height-24} {poly} {width-22},{height-24}"
    dots = "".join(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3.5" fill="#f59e0b"/>' for x, y in coords)
    return (
        f'<svg viewBox="0 0 {width} {height}" role="img" aria-label="Water level chart">'
        f'<rect width="{width}" height="{height}" fill="#fffaf0"/>'
        f'<line x1="22" y1="{height-24}" x2="{width-22}" y2="{height-24}" stroke="#f1d7ad"/>'
        f'<polygon points="{area}" fill="#fef3c7"/>'
        f'<polyline points="{poly}" fill="none" stroke="#b91c1c" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>'
        f"{dots}"
        f'<text x="22" y="18" font-size="12" fill="#7f1d1d">Water level: {min_v:.2f} to {max_v:.2f} m</text>'
        "</svg>"
    )


def _render_files(files: dict, token_q: str) -> str:
    rows = []
    for kind, info in files.items():
        exists = bool(info["exists"])
        status = "Ready" if exists else "Missing"
        status_class = "ok" if exists else "bad"
        modified = info.get("modified") or "No file"
        size = info.get("size_label") or "0 B"
        rows.append(
            f"""
            <tr>
              <td>
                <strong>{escape(info["label"])}</strong>
                <span>{escape(DATA_FILES[kind][0].name)}</span>
              </td>
              <td><mark class="{status_class}">{status}</mark></td>
              <td>{escape(size)}</td>
              <td>{escape(modified)}</td>
              <td class="actions">
                <a class="icon-link" href="/admin/download/{escape(kind)}?token={token_q}" aria-label="Download {escape(info["label"])}">Download</a>
                <form action="/admin/upload/{escape(kind)}?token={token_q}" method="post" enctype="multipart/form-data">
                  <input type="file" name="file" accept=".xlsx" required>
                  <button type="submit">Upload</button>
                </form>
              </td>
            </tr>
            """
        )
    return "".join(rows)


def _render_recent(events: list[dict]) -> str:
    if not events:
        return '<p class="empty">No recent LINE messages found.</p>'
    items = []
    for event in events:
        items.append(
            f"""
            <li>
              <time>{escape(event["time"])}</time>
              <strong>{escape(event["type"])}</strong>
              <span>{escape(event["date"])}</span>
              <p>{escape(event["text"])}</p>
            </li>
            """
        )
    return "<ol class=\"timeline\">" + "".join(items) + "</ol>"


def _render_flow(data: dict) -> str:
    env = data["env"]
    line_ok = env["line_secret"] and env["line_token"]
    db_ok = data["db_ok"]
    admin_ok = env["admin_token"]
    cron_paused = not env["weekly_enabled"]
    nodes = [
        ("LINE", "Messaging API", line_ok, False),
        ("FastAPI", "Webhook + admin", admin_ok, False),
        ("Supabase", "Reports + images", db_ok, False),
        ("Railway", env["railway_name"], True, False),
        ("Weekly", f'{env["weekly_schedule"]} {env["weekly_format"]}', env["weekly_enabled"], cron_paused),
    ]
    html_nodes = []
    for title, note, ok, paused in nodes:
        cls = _status_class(ok, paused)
        html_nodes.append(
            f"""
            <div class="flow-node {cls}">
              <i></i>
              <strong>{escape(title)}</strong>
              <span>{escape(note)}</span>
              <b>{escape(_status_word(ok, paused))}</b>
            </div>
            """
        )
    return "".join(html_nodes)


def _render_latest_daily(row: dict | None) -> str:
    if not row:
        return '<p class="empty">No daily report summary found.</p>'
    machinery = _equipment_count(row.get("equipment"))
    fields = [
        ("Date", row.get("work_date")),
        ("Status", row.get("report_status") or "draft"),
        ("Workers", row.get("total_workers") or 0),
        ("Engineers", row.get("engineers") or 0),
        ("Machinery", machinery),
        ("Foremen", row.get("foremen") or 0),
        ("Skilled", row.get("skilled_workers") or 0),
        ("Laborers", row.get("laborers") or 0),
        ("Water", "" if row.get("water_level") is None else f'{float(row.get("water_level")):.2f} m'),
        ("Weather", row.get("weather_morning") or "-"),
    ]
    cells = "".join(
        f"<div><span>{escape(label)}</span><strong>{escape(str(value))}</strong></div>" for label, value in fields
    )
    return f'<div class="fact-grid">{cells}</div>'


def _render_site_visual(data: dict) -> str:
    photos = data.get("recent_photos") or []
    hero_photo = next((p for p in photos if p.get("url")), None)
    image_html = ""
    if hero_photo:
        image_html = (
            f'<img class="site-photo" src="{escape(str(hero_photo["url"]))}" '
            f'alt="{escape(hero_photo["caption"])}">'
        )
    else:
        image_html = """
        <div class="site-placeholder" aria-label="Construction visualization">
          <span></span><span></span><span></span>
          <i></i>
        </div>
        """
    ctx = data["field_context"]
    return f"""
    <section class="site-stage">
      <div class="site-copy">
        <span class="eyebrow">Live Project Cockpit</span>
        <h2>{escape(PROJECT_DISPLAY_NAME)}</h2>
        <p>{escape(ctx["latest_text"])}</p>
      </div>
      <div class="site-visual">
        {image_html}
        <div class="live-chip">
          <b>Now</b>
          <span>{escape(ctx["latest_time"])}</span>
        </div>
        <div class="mini-card workers">
          <span>Workers</span>
          <strong>{escape(str(ctx["workers"]))}</strong>
        </div>
        <div class="mini-card engineers">
          <span>Engineers</span>
          <strong>{escape(str(ctx["engineers"]))}</strong>
        </div>
        <div class="mini-card machines">
          <span>Machines</span>
          <strong>{escape(str(ctx["equipment_count"]))}</strong>
        </div>
      </div>
    </section>
    """


def _render_today_activity(items: list[dict]) -> str:
    if not items:
        return '<p class="empty">No activity recorded for the current field date.</p>'
    html = []
    for item in items:
        html.append(
            f"""
            <li>
              <b>{escape(str(item["seq"]).zfill(2))}</b>
              <div>
                <span>{escape(item["type"])}</span>
                <p>{escape(item["text"])}</p>
              </div>
            </li>
            """
        )
    return '<ol class="activity-feed">' + "".join(html) + "</ol>"


def _render_finance_table(finance: dict) -> str:
    rows = list(finance.get("rows") or [])
    total = finance.get("total")
    display_rows = rows[:4]
    if total:
        display_rows.append(total)
    if not display_rows:
        return ""
    body = []
    for item in display_rows:
        label = str(item.get("label") or "-")
        row_class = ' class="total"' if label in {"รวม", "Total"} else ""
        body.append(
            f"""
            <tr{row_class}>
              <td>{escape(label)}</td>
              <td>{escape(_money(float(item.get("budget") or 0)))}</td>
              <td>{escape(_money(float(item.get("spent") or 0)))}</td>
              <td>{escape(_money(float(item.get("remaining") or 0)))}</td>
            </tr>
            """
        )
    return (
        '<table class="mini-table finance-table">'
        "<thead><tr><th>Year</th><th>Budget</th><th>Paid</th><th>Balance</th></tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table>"
    )


def _render_cockpit_cards(data: dict) -> str:
    ctx = data["field_context"]
    finance = data["financial_report"]
    progress = data["progress_report"]
    problem = data["problem_report"]
    finance_table = _render_finance_table(finance)
    return f"""
    <section class="cockpit-grid">
      <article class="glass-card assistant-card">
        <div class="card-head">
          <span class="eyebrow">Site Assistant</span>
          <b>{escape(ctx["work_date"] or "Live")}</b>
        </div>
        <p>{escape(ctx["latest_text"])}</p>
        <div class="assistant-actions">
          <span>Today activity</span>
          <span>Progress check</span>
          <span>Risk watch</span>
          <span>Report ready</span>
        </div>
      </article>

      <article class="glass-card finance-card">
        <div class="card-head">
          <span class="eyebrow">Financial Report</span>
          <b>{escape(finance["headline"])}</b>
        </div>
        <div class="progress-line"><i style="width:{int(finance["percent"])}%"></i></div>
        <dl class="compact-stats">
          <div><dt>Spent</dt><dd>{escape(finance["spent_label"])}</dd></div>
          <div><dt>Budget</dt><dd>{escape(finance["budget_label"])}</dd></div>
          <div><dt>Balance</dt><dd>{escape(finance["remaining_label"])}</dd></div>
        </dl>
        {finance_table}
        <small>{escape(finance["note"])}</small>
      </article>

      <article class="glass-card progress-card">
        <div class="card-head">
          <span class="eyebrow">Progress Report</span>
          <b>{escape(progress["headline"])}</b>
        </div>
        <div class="progress-line red"><i style="width:{int(progress["percent"])}%"></i></div>
        <dl class="compact-stats">
          <div><dt>Plan</dt><dd>{escape(_percent_label(progress["plan_percent"]))}</dd></div>
          <div><dt>Actual</dt><dd>{escape(_percent_label(progress["actual_percent"]))}</dd></div>
          <div><dt>Variance</dt><dd>{escape(_signed_percent_label(progress["variance_percent"]))}</dd></div>
        </dl>
        <small>{escape(progress["period_label"])} · Week {escape(str(progress["week_no"] or "-"))} · {escape(progress["note"])}</small>
      </article>

      <article class="glass-card problem-card">
        <div class="card-head">
          <span class="eyebrow">Problem Watch</span>
          <b>{escape(problem["status"])}</b>
        </div>
        <p>{escape(problem["latest"]["text"])}</p>
        <div class="problem-count">
          <strong>{escape(str(problem["count"]))}</strong>
          <span>recent warning keywords</span>
        </div>
      </article>
    </section>
    """


def _readiness_checks(data: dict) -> list[dict]:
    ctx = data["field_context"]
    latest = data.get("latest_daily") or {}
    progress = data["progress_report"]
    finance = data["financial_report"]
    problem = data["problem_report"]
    checks = [
        {
            "label": "Daily Report",
            "ok": bool(latest),
            "note": ctx["work_date"] or "No active daily report",
        },
        {
            "label": "Activities",
            "ok": bool(data["today_activities"]),
            "note": f"{len(data['today_activities'])} items on active date",
        },
        {
            "label": "Photo Evidence",
            "ok": int(ctx["photo_count"] or 0) > 0,
            "note": f"{ctx['photo_count']} photos in recent feed",
        },
        {
            "label": "Workforce",
            "ok": int(ctx["workers"] or 0) > 0,
            "note": f"{ctx['workers']} workers / {ctx['engineers']} engineers",
        },
        {
            "label": "Machinery",
            "ok": int(ctx["equipment_count"] or 0) > 0,
            "note": f"{ctx['equipment_count']} active units",
        },
        {
            "label": "Cost Source",
            "ok": bool(finance.get("connected")),
            "note": finance.get("note") or "Financial source pending",
        },
        {
            "label": "Schedule",
            "ok": float(progress.get("variance_percent") or 0) >= 0,
            "note": progress.get("headline") or "Progress source pending",
        },
        {
            "label": "Risk Signals",
            "ok": int(problem.get("count") or 0) == 0,
            "note": f"{problem.get('count', 0)} warning keywords",
            "warn": int(problem.get("count") or 0) > 0,
        },
    ]
    return checks


def _decision_items(data: dict) -> list[str]:
    ctx = data["field_context"]
    progress = data["progress_report"]
    problem = data["problem_report"]
    items = []
    if int(problem.get("count") or 0) > 0:
        items.append(f"Review {problem['count']} field issue signal(s) and assign an owner.")
    if not data["today_activities"]:
        items.append("Confirm today's activity log from the field team.")
    if int(ctx["photo_count"] or 0) == 0:
        items.append("Request field photo evidence for the current report cycle.")
    if int(ctx["workers"] or 0) == 0:
        items.append("Confirm workforce count before issuing the daily brief.")
    if float(progress.get("variance_percent") or 0) < 0:
        items.append("Prepare schedule recovery action because progress is behind plan.")
    if not items:
        items.append("No critical exception is flagged from the current dashboard data.")
    return items[:4]


def _render_admin_command_brief(data: dict) -> str:
    ctx = data["field_context"]
    progress = data["progress_report"]
    finance = data["financial_report"]
    problem = data["problem_report"]
    checks = _readiness_checks(data)
    decision_rows = "".join(f"<li>{escape(item)}</li>" for item in _decision_items(data))
    check_rows = []
    for check in checks:
        cls = "warn" if check.get("warn") else "ok" if check["ok"] else "bad"
        word = "Watch" if cls == "warn" else "Ready" if cls == "ok" else "Missing"
        check_rows.append(
            f"""
            <div class="readiness-row {cls}">
              <div><strong>{escape(check["label"])}</strong><span>{escape(check["note"])}</span></div>
              <b>{word}</b>
            </div>
            """
        )
    return f"""
    <section class="command-brief">
      <article class="decision-panel">
        <span class="eyebrow">Today's Decision</span>
        <h2>{escape(ctx["work_date"] or "Live Project Control")}</h2>
        <ol>{decision_rows}</ol>
      </article>
      <article class="readiness-panel">
        <div class="panel-head">
          <div>
            <span class="eyebrow">Field Readiness</span>
            <h2>Exception Check</h2>
          </div>
          <mark class="ok">{escape(problem["status"])}</mark>
        </div>
        <div class="readiness-list">{''.join(check_rows)}</div>
      </article>
      <article class="control-stat schedule">
        <span>Schedule Health</span>
        <strong>{escape(_signed_percent_label(progress["variance_percent"]))}</strong>
        <small>{escape(progress["period_label"])}</small>
      </article>
      <article class="control-stat cost">
        <span>Cost Health</span>
        <strong>{escape(str(finance["headline"]).split()[0])}</strong>
        <small>{escape(finance["spent_label"])} paid</small>
      </article>
      <article class="control-stat resource">
        <span>Resources</span>
        <strong>{escape(str(ctx["workers"]))} / {escape(str(ctx["equipment_count"]))}</strong>
        <small>Workers / machines</small>
      </article>
    </section>
    """


def _issue_status_label(status: object) -> str:
    text = str(status or "open").replace("_", " ").strip().title()
    return text or "Open"


def _issue_badge_class(row: dict) -> str:
    status = str(row.get("status") or "open").lower()
    impact = str(row.get("impact") or "medium").lower()
    due = _parse_date(row.get("due_date"))
    today = _now_bkk().date()
    if status in {"resolved", "closed"}:
        return "ok"
    if due and due < today:
        return "bad"
    if impact in {"high", "critical"} or status == "waiting":
        return "warn"
    return "open"


def _render_problem_board(data: dict, token_q: str) -> str:
    board = data.get("issue_board") or {}
    issues = board.get("open_issues") or []
    comments_by_issue = board.get("comments_by_issue") or {}
    setup_note = ""
    if data.get("issues_error"):
        setup_note = (
            '<div class="error">Problem board table is not connected yet. '
            'Run setup_problem_board.sql in Supabase SQL Editor.</div>'
        )

    issue_cards = []
    for row in issues[:8]:
        issue_id = int(row.get("id") or 0)
        comments = comments_by_issue.get(issue_id, [])
        latest_comment = comments[0] if comments else None
        due = str(row.get("due_date") or "No due date")
        owner = str(row.get("owner") or "")
        cls = _issue_badge_class(row)
        latest_comment_html = ""
        if latest_comment:
            latest_comment_html = (
                f'<p class="issue-comment">{escape(_short(latest_comment.get("comment"), 100))}</p>'
            )
        issue_cards.append(
            f"""
            <article class="issue-card {cls}">
              <div class="issue-main">
                <div>
                  <span class="issue-meta">{escape(str(row.get("area") or "Project"))} · {escape(due)}</span>
                  <h3>{escape(_short(row.get("title") or row.get("description") or "Untitled issue", 92))}</h3>
                  <p>{escape(_short(row.get("next_action") or row.get("description") or "No next action recorded.", 130))}</p>
                  {latest_comment_html}
                </div>
                <mark class="{cls}">{escape(_issue_status_label(row.get("status")))}</mark>
              </div>
              <form class="issue-update" action="/admin/issues/{issue_id}/update?token={token_q}" method="post">
                <input type="text" name="owner" value="{escape(owner)}" placeholder="Owner" aria-label="Owner">
                <input type="date" name="due_date" value="{escape(str(row.get("due_date") or ""))}" aria-label="Due date">
                <select name="status" aria-label="Status">
                  {_status_options(str(row.get("status") or "open"))}
                </select>
                <input type="text" name="next_action" value="{escape(str(row.get("next_action") or ""))}" placeholder="Next action">
                <button type="submit">Update</button>
              </form>
              <form class="issue-comment-form" action="/admin/issues/{issue_id}/comment?token={token_q}" method="post">
                <input type="text" name="comment" placeholder="Add discussion note">
                <button type="submit">Comment</button>
              </form>
            </article>
            """
        )

    if not issue_cards:
        issue_cards.append(
            """
            <article class="issue-card empty-board">
              <h3>No open project issues</h3>
              <p>LINE issue reports and web-created decisions will appear here.</p>
            </article>
            """
        )

    return f"""
    <section class="section problem-board">
      <div class="split-head">
        <div>
          <h2>Problem & Decision Board</h2>
          <p class="section-lead">LINE reports create issues automatically; admin uses this board to assign, discuss and close them.</p>
        </div>
        <span class="pill">{escape(str(board.get("open_count", 0)))} open · {escape(str(board.get("overdue_count", 0)))} overdue</span>
      </div>
      {setup_note}
      <form class="issue-create" action="/admin/issues/create?token={token_q}" method="post">
        <input type="text" name="title" placeholder="Problem / decision title" required>
        <input type="text" name="area" placeholder="Area">
        <input type="text" name="owner" placeholder="Owner">
        <input type="date" name="due_date" aria-label="Due date">
        <select name="impact" aria-label="Impact">
          <option value="medium">Medium</option>
          <option value="low">Low</option>
          <option value="high">High</option>
          <option value="critical">Critical</option>
        </select>
        <input type="text" name="next_action" placeholder="Next action">
        <button class="primary" type="submit">Add Issue</button>
      </form>
      <div class="issue-grid">{''.join(issue_cards)}</div>
    </section>
    """


def _status_options(current: str) -> str:
    statuses = [
        ("open", "Open"),
        ("in_progress", "In Progress"),
        ("waiting", "Waiting"),
        ("resolved", "Resolved"),
        ("closed", "Closed"),
    ]
    return "".join(
        f'<option value="{value}"{" selected" if value == current else ""}>{label}</option>'
        for value, label in statuses
    )


def _render_photo_strip(photos: list[dict]) -> str:
    if not photos:
        return """
        <div class="photo-strip empty-photos">
          <div></div><div></div><div></div>
        </div>
        """
    cards = []
    for photo in photos[:4]:
        if photo.get("url"):
            visual = f'<img src="{escape(str(photo["url"]))}" alt="{escape(photo["caption"])}">'
        else:
            visual = "<div></div>"
        cards.append(
            f"""
            <figure>
              {visual}
              <figcaption>{escape(photo["caption"])}</figcaption>
            </figure>
            """
        )
    return '<div class="photo-strip">' + "".join(cards) + "</div>"


def _page_css() -> str:
    return """
    :root {
      --ink: #171717;
      --muted: #6d7176;
      --line: #deded9;
      --page: #d9d9d6;
      --panel: #ffffff;
      --soft: #f1f1ee;
      --red: #b51f1a;
      --red-dark: #5b1110;
      --gold: #c79221;
      --gold-soft: #f4dfad;
      --green: #10a36f;
      --steel: #242424;
      --shadow: 0 18px 50px rgba(22, 22, 22, .10);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      color: var(--ink);
      background: var(--page);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background:
        linear-gradient(180deg, #ededeb 0, var(--page) 420px, #d8d8d5 100%);
    }
    a { color: inherit; }
    .shell { min-height: 100vh; }
    .topbar {
      position: sticky;
      top: 0;
      z-index: 5;
      display: flex;
      justify-content: space-between;
      gap: 18px;
      align-items: center;
      padding: 16px clamp(18px, 3vw, 42px);
      border-bottom: 1px solid rgba(23, 23, 23, .08);
      background: rgba(240, 240, 237, .88);
      backdrop-filter: blur(12px);
    }
    .brand { display: flex; gap: 12px; align-items: center; min-width: 0; color: var(--ink); }
    .brand-mark {
      width: 42px;
      height: 42px;
      border-radius: 8px;
      background:
        linear-gradient(135deg, var(--red) 0 48%, transparent 49%),
        linear-gradient(135deg, var(--gold) 0 68%, var(--ink) 69% 100%);
      box-shadow: inset 0 0 0 1px rgba(255, 255, 255, .45);
      flex: 0 0 auto;
    }
    .brand h1 { font-size: 18px; line-height: 1.15; margin: 0; letter-spacing: 0; }
    .brand span { display: block; color: var(--muted); font-size: 12px; margin-top: 3px; max-width: 720px; overflow-wrap: anywhere; }
    .top-actions { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; justify-content: flex-end; }
    .pill {
      display: inline-flex;
      gap: 8px;
      align-items: center;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 7px 11px;
      background: rgba(255, 255, 255, .72);
      color: var(--ink);
      font-size: 12px;
      white-space: nowrap;
    }
    .dot { width: 8px; height: 8px; border-radius: 50%; background: var(--green); display: inline-block; }
    .dot.warn { background: var(--gold); }
    .dot.bad { background: var(--red); }
    main { padding: 24px clamp(18px, 3vw, 42px) 48px; }
    .command-brief {
      display: grid;
      grid-template-columns: minmax(320px, 1.2fr) minmax(320px, 1fr) repeat(3, minmax(150px, .55fr));
      gap: 14px;
      margin-bottom: 18px;
      align-items: stretch;
    }
    .decision-panel, .readiness-panel, .control-stat {
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      min-width: 0;
    }
    .decision-panel {
      background: #181818;
      color: #fff;
      padding: 22px;
      display: flex;
      flex-direction: column;
      justify-content: space-between;
      gap: 18px;
    }
    .decision-panel .eyebrow { color: var(--gold-soft); }
    .decision-panel h2 { margin: 0; font-size: 26px; line-height: 1.15; letter-spacing: 0; }
    .decision-panel ol { margin: 0; padding: 0; list-style: none; display: grid; gap: 10px; }
    .decision-panel li {
      padding-left: 16px;
      border-left: 3px solid var(--gold);
      color: rgba(255,255,255,.86);
      line-height: 1.42;
      font-size: 14px;
    }
    .readiness-panel { background: rgba(255,255,255,.88); padding: 18px; }
    .panel-head { display: flex; justify-content: space-between; align-items: flex-start; gap: 14px; margin-bottom: 10px; }
    .panel-head h2 { margin: 0; font-size: 18px; letter-spacing: 0; }
    .readiness-list { display: grid; gap: 8px; }
    .readiness-row {
      display: grid;
      grid-template-columns: 1fr 72px;
      gap: 10px;
      align-items: center;
      padding: 9px 0;
      border-bottom: 1px solid rgba(23,23,23,.08);
    }
    .readiness-row:last-child { border-bottom: 0; }
    .readiness-row strong { display: block; font-size: 13px; }
    .readiness-row span { display: block; color: var(--muted); font-size: 12px; margin-top: 2px; line-height: 1.35; }
    .readiness-row b {
      justify-self: end;
      border-radius: 999px;
      padding: 5px 8px;
      background: #edf7f2;
      color: #0f6d50;
      font-size: 11px;
    }
    .readiness-row.warn b { background: #fff4d5; color: #7c4b06; }
    .readiness-row.bad b { background: #fee2e2; color: #991b1b; }
    .control-stat {
      background: var(--panel);
      padding: 16px;
      display: flex;
      flex-direction: column;
      justify-content: space-between;
      border-top: 4px solid var(--red);
    }
    .control-stat.cost, .control-stat.resource { border-top-color: var(--gold); }
    .control-stat span { color: var(--muted); font-size: 12px; }
    .control-stat strong { display: block; font-size: 28px; line-height: 1; margin: 14px 0 8px; }
    .control-stat small { color: var(--muted); line-height: 1.35; }
    .site-stage {
      display: grid;
      grid-template-columns: minmax(280px, 1fr) minmax(360px, .9fr);
      gap: 18px;
      align-items: stretch;
      margin-bottom: 18px;
      min-height: 430px;
      padding: 26px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #eeeeeb;
      box-shadow: var(--shadow);
      overflow: hidden;
    }
    .site-copy { display: flex; flex-direction: column; justify-content: center; max-width: 760px; }
    .eyebrow {
      display: block;
      color: var(--red);
      font-size: 12px;
      text-transform: uppercase;
      font-weight: 800;
      letter-spacing: 0;
      margin-bottom: 10px;
    }
    .site-copy h2 {
      margin: 0;
      font-size: 32px;
      line-height: 1.24;
      letter-spacing: 0;
    }
    .site-copy p { max-width: 620px; color: var(--muted); font-size: 15px; line-height: 1.6; margin: 18px 0 0; }
    .site-visual { position: relative; min-height: 350px; border-radius: 8px; background: #e5e5e1; overflow: hidden; }
    .site-photo { width: 100%; height: 100%; object-fit: cover; display: block; filter: saturate(.9); }
    .site-placeholder {
      position: absolute;
      inset: 0;
      background:
        linear-gradient(155deg, transparent 0 48%, rgba(181, 31, 26, .16) 49% 51%, transparent 52%),
        linear-gradient(180deg, #eeeeeb, #d5d5d1);
    }
    .site-placeholder span { position: absolute; background: rgba(255,255,255,.62); border: 1px solid rgba(0,0,0,.04); border-radius: 8px; }
    .site-placeholder span:nth-child(1) { width: 62%; height: 16%; left: 16%; top: 18%; transform: rotate(-8deg); }
    .site-placeholder span:nth-child(2) { width: 52%; height: 12%; right: 8%; top: 45%; transform: rotate(8deg); }
    .site-placeholder span:nth-child(3) { width: 76%; height: 18%; left: 10%; bottom: 16%; transform: rotate(-3deg); }
    .site-placeholder i { position: absolute; left: 22%; bottom: 24%; width: 44%; height: 6px; background: var(--ink); border-radius: 999px; opacity: .82; }
    .live-chip, .mini-card {
      position: absolute;
      border: 1px solid rgba(255,255,255,.74);
      background: rgba(255,255,255,.78);
      backdrop-filter: blur(12px);
      border-radius: 8px;
      box-shadow: 0 12px 30px rgba(23,23,23,.12);
    }
    .live-chip { top: 18px; right: 18px; padding: 12px 14px; min-width: 170px; }
    .live-chip b, .mini-card span { display: block; font-size: 11px; color: var(--muted); }
    .live-chip span { display: block; margin-top: 4px; font-weight: 750; }
    .mini-card { padding: 13px 14px; min-width: 96px; }
    .mini-card strong { display: block; font-size: 28px; line-height: 1; margin-top: 6px; }
    .mini-card.workers { left: 18px; bottom: 18px; }
    .mini-card.engineers { left: 50%; bottom: 18px; transform: translateX(-50%); }
    .mini-card.machines { right: 18px; bottom: 18px; }
    .cockpit-grid {
      display: grid;
      grid-template-columns: 1.35fr repeat(3, minmax(180px, 1fr));
      gap: 14px;
      margin-bottom: 18px;
    }
    .glass-card {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: rgba(255,255,255,.82);
      box-shadow: var(--shadow);
      padding: 18px;
      min-width: 0;
    }
    .card-head { display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; margin-bottom: 12px; }
    .card-head b { font-size: 13px; color: var(--ink); text-align: right; }
    .assistant-card p, .problem-card p { color: var(--muted); line-height: 1.55; margin: 0 0 14px; }
    .assistant-actions { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }
    .assistant-actions span {
      border-radius: 8px;
      background: var(--soft);
      padding: 10px;
      font-size: 12px;
      font-weight: 700;
    }
    .progress-line { height: 8px; border-radius: 999px; background: #e6e3dc; overflow: hidden; margin: 12px 0 14px; }
    .progress-line i { display: block; height: 100%; border-radius: inherit; background: var(--gold); }
    .progress-line.red i { background: var(--red); }
    .compact-stats { display: grid; grid-template-columns: repeat(3, minmax(0,1fr)); gap: 8px; margin: 0 0 12px; }
    .compact-stats div { background: var(--soft); border-radius: 8px; padding: 9px; }
    .compact-stats dt { color: var(--muted); font-size: 11px; }
    .compact-stats dd { margin: 4px 0 0; font-weight: 800; }
    .mini-table { width: 100%; border-collapse: collapse; margin: 2px 0 12px; font-size: 11px; }
    .mini-table th, .mini-table td { padding: 6px 4px; border-bottom: 1px solid rgba(23,23,23,.08); text-align: right; white-space: nowrap; }
    .mini-table th:first-child, .mini-table td:first-child { text-align: left; white-space: normal; }
    .mini-table th { color: var(--muted); font-weight: 750; }
    .mini-table .total td { color: var(--red); font-weight: 850; border-bottom: 0; }
    .glass-card small { color: var(--muted); line-height: 1.35; display: block; }
    .problem-count { display: flex; align-items: end; gap: 10px; }
    .problem-count strong { font-size: 36px; line-height: 1; color: var(--red); }
    .problem-count span { color: var(--muted); font-size: 12px; }
    .dashboard-grid {
      display: grid;
      grid-template-columns: minmax(280px, .8fr) minmax(320px, 1.2fr);
      gap: 18px;
      margin-bottom: 18px;
      align-items: start;
    }
    .summary {
      display: grid;
      grid-template-columns: minmax(260px, 1.2fr) minmax(280px, 2fr);
      gap: 18px;
      align-items: stretch;
    }
    .hero-panel, .chart-panel, .section {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .hero-panel {
      padding: 22px;
      display: grid;
      grid-template-rows: auto 1fr auto;
      gap: 20px;
      overflow: hidden;
      min-height: 300px;
    }
    .health-ring {
      width: min(220px, 52vw);
      aspect-ratio: 1;
      border-radius: 50%;
      margin: 4px auto;
      display: grid;
      place-items: center;
      background:
        radial-gradient(circle at center, #fff 0 57%, transparent 58%),
        conic-gradient(var(--red) calc(var(--score) * 1%), #e5e2db 0);
      border: 1px solid var(--line);
    }
    .health-ring strong { display: block; font-size: 46px; line-height: 1; }
    .health-ring span { color: var(--muted); font-size: 12px; display: block; text-align: center; margin-top: 6px; }
    .hero-panel h2, .chart-panel h2, .section h2 {
      margin: 0;
      font-size: 14px;
      letter-spacing: 0;
      text-transform: uppercase;
      color: var(--ink);
    }
    .runtime { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
    .runtime div { border-top: 1px solid var(--line); padding-top: 10px; min-width: 0; }
    .runtime span, .fact-grid span { display: block; color: var(--muted); font-size: 12px; }
    .runtime strong, .fact-grid strong { display: block; margin-top: 4px; overflow-wrap: anywhere; }
    .chart-panel { padding: 18px; overflow: hidden; }
    .chart-panel svg { width: 100%; height: auto; display: block; margin-top: 10px; }
    .metrics {
      display: grid;
      grid-template-columns: repeat(6, minmax(140px, 1fr));
      gap: 12px;
      margin: 18px 0;
    }
    .metric {
      background: var(--panel);
      border: 1px solid var(--line);
      border-top: 4px solid var(--accent);
      border-radius: 8px;
      padding: 14px;
      min-height: 112px;
      box-shadow: none;
    }
    .metric-key { color: var(--muted); font-size: 12px; display: block; }
    .metric strong { display: block; font-size: 30px; margin: 8px 0 5px; letter-spacing: 0; }
    .metric small { color: var(--muted); line-height: 1.35; display: block; }
    .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 18px; align-items: start; }
    .section { padding: 18px; min-width: 0; }
    .section-lead { color: var(--muted); margin: 6px 0 0; line-height: 1.4; font-size: 13px; }
    .problem-board { margin: 18px 0; }
    .issue-create {
      display: grid;
      grid-template-columns: minmax(220px, 1.4fr) repeat(4, minmax(110px, .7fr)) minmax(180px, 1fr) auto;
      gap: 8px;
      align-items: center;
      margin-top: 16px;
      padding: 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--soft);
    }
    .issue-grid { display: grid; gap: 10px; margin-top: 14px; }
    .issue-card {
      border: 1px solid var(--line);
      border-left: 4px solid var(--gold);
      border-radius: 8px;
      background: #fff;
      padding: 14px;
      min-width: 0;
    }
    .issue-card.bad { border-left-color: var(--red); background: #fffafa; }
    .issue-card.warn { border-left-color: var(--gold); }
    .issue-card.ok { border-left-color: var(--green); }
    .issue-main {
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 14px;
      align-items: start;
      margin-bottom: 12px;
    }
    .issue-meta { display: block; color: var(--muted); font-size: 12px; margin-bottom: 5px; }
    .issue-card h3 { margin: 0; font-size: 17px; line-height: 1.25; }
    .issue-card p { margin: 7px 0 0; color: var(--muted); line-height: 1.45; }
    .issue-comment { color: var(--ink); background: #f7f4ed; padding: 8px 10px; border-radius: 8px; }
    .issue-update, .issue-comment-form {
      display: grid;
      grid-template-columns: 150px 150px 140px minmax(180px, 1fr) auto;
      gap: 8px;
      align-items: center;
    }
    .issue-comment-form { grid-template-columns: 1fr auto; margin-top: 8px; }
    .empty-board { text-align: left; }
    .flow {
      display: grid;
      grid-template-columns: repeat(5, minmax(140px, 1fr));
      gap: 12px;
      margin-top: 14px;
    }
    .flow-node {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px;
      min-height: 136px;
      position: relative;
      background: #fff;
    }
    .flow-node i {
      width: 34px;
      height: 34px;
      display: block;
      border-radius: 8px;
      background: #fee2e2;
      margin-bottom: 12px;
      position: relative;
    }
    .flow-node i:after {
      content: "";
      position: absolute;
      inset: 9px;
      border: 2px solid var(--red);
      border-radius: 3px;
    }
    .flow-node.ok i { background: #fef3c7; }
    .flow-node.ok i:after { border-color: var(--gold); }
    .flow-node.warn i { background: #fef3c7; }
    .flow-node.warn i:after { border-color: var(--gold); }
    .flow-node.bad i { background: #fee2e2; }
    .flow-node.bad i:after { border-color: var(--red); }
    .flow-node strong { display: block; font-size: 15px; }
    .flow-node span { display: block; color: var(--muted); font-size: 12px; margin: 5px 0 12px; min-height: 32px; overflow-wrap: anywhere; }
    .flow-node b { font-size: 12px; color: var(--red); }
    .bar-row {
      display: grid;
      grid-template-columns: minmax(96px, 150px) 1fr 38px;
      gap: 10px;
      align-items: center;
      margin: 12px 0;
      font-size: 13px;
    }
    .bar-row span { color: var(--steel); overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .bar-track { height: 11px; background: #e6e3dc; border-radius: 999px; overflow: hidden; }
    .bar-track i { display: block; height: 100%; border-radius: 999px; background: var(--red); }
    .bar-row b { text-align: right; }
    table { width: 100%; border-collapse: collapse; margin-top: 14px; font-size: 13px; }
    th { text-align: left; color: var(--muted); font-weight: 700; border-bottom: 1px solid var(--line); padding: 10px 8px; }
    td { border-bottom: 1px solid #f3e3c8; padding: 11px 8px; vertical-align: middle; }
    td strong { display: block; }
    td span { color: var(--muted); font-size: 12px; display: block; margin-top: 2px; }
    mark {
      border-radius: 999px;
      padding: 5px 9px;
      font-size: 12px;
      color: var(--steel);
      background: #f3e8d1;
    }
    mark.ok { background: #fef3c7; color: #7c2d12; }
    mark.bad { background: #fee2e2; color: #991b1b; }
    .actions { display: flex; gap: 8px; align-items: center; flex-wrap: wrap; }
    .actions form { display: flex; gap: 8px; align-items: center; flex-wrap: wrap; }
    input[type="text"], input[type="date"], select {
      width: 100%;
      min-height: 34px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
      color: var(--ink);
      padding: 7px 10px;
      font: inherit;
      font-size: 13px;
      min-width: 0;
    }
    input[type="file"] {
      max-width: 210px;
      font-size: 12px;
      color: var(--ink);
    }
    button, .icon-link {
      appearance: none;
      border: 1px solid #e4c993;
      background: #fff;
      color: var(--ink);
      min-height: 34px;
      padding: 7px 12px;
      border-radius: 8px;
      font-weight: 650;
      font-size: 13px;
      text-decoration: none;
      cursor: pointer;
      white-space: nowrap;
    }
    button.primary {
      border-color: var(--ink);
      background: var(--ink);
      color: #fff;
    }
    button:hover, .icon-link:hover { border-color: var(--gold); box-shadow: 0 0 0 3px rgba(245, 158, 11, .16); }
    .timeline {
      list-style: none;
      padding: 0;
      margin: 14px 0 0;
      display: grid;
      gap: 10px;
    }
    .timeline li {
      border-left: 3px solid var(--gold);
      padding: 2px 0 4px 12px;
    }
    .timeline time { color: var(--muted); font-size: 12px; display: block; }
    .timeline strong { display: inline-block; margin-top: 2px; }
    .timeline span { color: var(--muted); font-size: 12px; margin-left: 8px; }
    .timeline p { margin: 5px 0 0; color: var(--steel); line-height: 1.35; }
    .fact-grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
      margin-top: 14px;
    }
    .fact-grid div {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      background: #fffaf0;
      min-width: 0;
    }
    .activity-feed { list-style: none; padding: 0; margin: 14px 0 0; display: grid; gap: 10px; }
    .activity-feed li { display: grid; grid-template-columns: 38px 1fr; gap: 10px; align-items: start; }
    .activity-feed b { width: 34px; height: 34px; display: grid; place-items: center; border-radius: 8px; background: var(--ink); color: #fff; font-size: 12px; }
    .activity-feed span { color: var(--red); font-size: 12px; font-weight: 800; }
    .activity-feed p { margin: 3px 0 0; color: var(--ink); line-height: 1.4; }
    .photo-strip { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 10px; margin-top: 14px; }
    .photo-strip figure { margin: 0; min-width: 0; }
    .photo-strip img, .photo-strip figure > div, .empty-photos div {
      display: block;
      width: 100%;
      aspect-ratio: 1.15;
      object-fit: cover;
      background: linear-gradient(135deg, #eeeeeb, #cfcfca);
      border-radius: 8px;
      border: 1px solid var(--line);
    }
    .photo-strip figcaption { color: var(--muted); font-size: 11px; margin-top: 6px; line-height: 1.35; }
    .split-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 2px;
    }
    .error {
      margin-top: 12px;
      border: 1px solid #fecaca;
      background: #fff7f7;
      color: #991b1b;
      border-radius: 8px;
      padding: 10px 12px;
      font-size: 13px;
      overflow-wrap: anywhere;
    }
    .empty { color: var(--muted); margin: 14px 0 0; }
    .footnote { color: var(--muted); font-size: 12px; margin-top: 14px; }
    @media (max-width: 1120px) {
      .command-brief { grid-template-columns: 1fr 1fr; }
      .control-stat { min-height: 132px; }
      .metrics { grid-template-columns: repeat(3, 1fr); }
      .summary, .grid-2, .site-stage, .dashboard-grid { grid-template-columns: 1fr; }
      .cockpit-grid { grid-template-columns: 1fr 1fr; }
      .flow { grid-template-columns: repeat(2, minmax(150px, 1fr)); }
    }
    @media (max-width: 720px) {
      .topbar { align-items: flex-start; flex-direction: column; }
      .command-brief { grid-template-columns: 1fr; }
      .site-stage { padding: 18px; min-height: 0; }
      .site-copy h2 { font-size: 22px; }
      .site-visual { min-height: 300px; }
      .cockpit-grid { grid-template-columns: 1fr; }
      .metrics { grid-template-columns: 1fr 1fr; }
      .runtime, .fact-grid { grid-template-columns: 1fr; }
      .flow { grid-template-columns: 1fr; }
      .photo-strip { grid-template-columns: 1fr 1fr; }
      .bar-row { grid-template-columns: 1fr 54px; }
      .bar-row .bar-track { grid-column: 1 / -1; order: 3; }
      .issue-create, .issue-update { grid-template-columns: 1fr 1fr; }
      .issue-create button, .issue-update button, .issue-comment-form { grid-column: 1 / -1; }
      table, thead, tbody, tr, th, td { display: block; }
      thead { display: none; }
      tr { border: 1px solid var(--line); border-radius: 8px; margin: 12px 0; background: #fff; }
      td { border-bottom: 0; }
      .actions { align-items: stretch; }
    }
    @media (max-width: 480px) {
      .metrics { grid-template-columns: 1fr; }
      main { padding-left: 14px; padding-right: 14px; }
      .brand h1 { font-size: 16px; }
      .compact-stats { grid-template-columns: 1fr; }
      .mini-card { min-width: 84px; padding: 10px; }
      .mini-card strong { font-size: 24px; }
      .issue-create, .issue-update, .issue-comment-form { grid-template-columns: 1fr; }
      .issue-main { grid-template-columns: 1fr; }
    }
    """


def _public_css() -> str:
    return """
    :root {
      --ink: #141414;
      --muted: #73777c;
      --page: #e4e4e1;
      --panel: #f8f8f5;
      --white: #ffffff;
      --line: rgba(20,20,20,.11);
      --red: #b51f1a;
      --red-dark: #5a1210;
      --gold: #c89325;
      --gold-soft: #f2d27e;
      --green: #12966e;
      --shadow: 0 24px 70px rgba(20,20,20,.14);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      color: var(--ink);
      background: var(--page);
    }
    * { box-sizing: border-box; }
    body { margin: 0; background: var(--page); }
    .public-shell { min-height: 100vh; }
    .public-nav {
      position: sticky;
      top: 0;
      z-index: 5;
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 18px;
      padding: 16px clamp(18px, 4vw, 54px);
      background: rgba(231,231,228,.86);
      backdrop-filter: blur(14px);
      border-bottom: 1px solid var(--line);
    }
    .public-brand { display: flex; align-items: center; gap: 12px; min-width: 0; }
    .public-mark {
      width: 38px;
      height: 38px;
      border-radius: 8px;
      background:
        linear-gradient(145deg, var(--red) 0 44%, transparent 45%),
        linear-gradient(145deg, var(--gold) 0 74%, #171717 75%);
      box-shadow: inset 0 0 0 1px rgba(255,255,255,.54);
      flex: 0 0 auto;
    }
    .public-brand b { display: block; font-size: 14px; letter-spacing: 0; }
    .public-brand span { display: block; color: var(--muted); font-size: 12px; margin-top: 2px; }
    .public-pills { display: flex; flex-wrap: wrap; justify-content: flex-end; gap: 9px; }
    .public-links { display: flex; gap: 4px; align-items: center; flex-wrap: wrap; }
    .public-links a {
      color: var(--ink);
      text-decoration: none;
      font-size: 12px;
      font-weight: 750;
      padding: 8px 10px;
      border-radius: 999px;
    }
    .public-links a:hover { background: rgba(255,255,255,.7); }
    .public-pill {
      display: inline-flex;
      align-items: center;
      gap: 7px;
      border: 1px solid var(--line);
      border-radius: 999px;
      background: rgba(255,255,255,.64);
      padding: 8px 11px;
      color: var(--ink);
      font-size: 12px;
      white-space: nowrap;
    }
    .public-dot { width: 8px; height: 8px; border-radius: 50%; background: var(--green); display: inline-block; }
    main { padding: 24px clamp(18px, 4vw, 54px) 54px; }
    .public-hero {
      min-height: min(680px, calc(100vh - 112px));
      display: grid;
      grid-template-columns: minmax(320px, .92fr) minmax(420px, 1.08fr);
      gap: 20px;
      align-items: stretch;
      margin-bottom: 20px;
    }
    .public-hero-copy {
      border: 1px solid var(--line);
      border-radius: 8px;
      background:
        linear-gradient(180deg, rgba(255,255,255,.82), rgba(248,248,245,.96)),
        var(--panel);
      box-shadow: var(--shadow);
      padding: clamp(24px, 4vw, 46px);
      display: flex;
      flex-direction: column;
      justify-content: space-between;
      gap: 28px;
      min-width: 0;
    }
    .public-eyebrow {
      display: block;
      color: var(--red);
      font-size: 12px;
      font-weight: 850;
      letter-spacing: 0;
      text-transform: uppercase;
      margin-bottom: 12px;
    }
    .public-hero h1 {
      margin: 0;
      font-size: clamp(30px, 4.2vw, 58px);
      line-height: 1.08;
      letter-spacing: 0;
      max-width: 860px;
    }
    .public-lead {
      color: var(--muted);
      font-size: 16px;
      line-height: 1.65;
      max-width: 680px;
      margin: 18px 0 0;
    }
    .public-source {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 20px;
    }
    .public-source span {
      display: inline-flex;
      align-items: center;
      border: 1px solid var(--line);
      border-radius: 999px;
      background: rgba(255,255,255,.72);
      color: var(--muted);
      font-size: 12px;
      padding: 7px 10px;
    }
    .public-hero-bottom {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 12px;
    }
    .public-signal {
      background: #ecece8;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px;
      min-width: 0;
    }
    .public-signal span { display: block; color: var(--muted); font-size: 12px; }
    .public-signal strong { display: block; margin-top: 6px; font-size: 20px; line-height: 1.22; overflow-wrap: anywhere; }
    .public-exec-band {
      display: grid;
      grid-template-columns: 1.2fr repeat(3, minmax(160px, .7fr));
      gap: 12px;
      margin-bottom: 20px;
    }
    .public-exec-band article {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: rgba(255,255,255,.82);
      padding: 18px;
      min-width: 0;
    }
    .public-exec-band article:first-child {
      background: #171717;
      color: #fff;
    }
    .public-exec-band span { display: block; color: var(--muted); font-size: 12px; }
    .public-exec-band article:first-child span { color: var(--gold-soft); }
    .public-exec-band strong { display: block; font-size: 24px; margin-top: 8px; line-height: 1.15; overflow-wrap: anywhere; }
    .public-exec-band p { margin: 8px 0 0; color: rgba(255,255,255,.75); line-height: 1.45; }
    .public-visual {
      border-radius: 8px;
      border: 1px solid var(--line);
      box-shadow: var(--shadow);
      background: #d9d9d5;
      min-height: 520px;
      overflow: hidden;
      position: relative;
    }
    .public-visual img { width: 100%; height: 100%; object-fit: cover; display: block; filter: saturate(.92) contrast(.98); }
    .public-visual-fallback {
      position: absolute;
      inset: 0;
      background:
        linear-gradient(145deg, rgba(181,31,26,.16), transparent 38%),
        linear-gradient(18deg, rgba(200,147,37,.24), transparent 48%),
        #dededb;
    }
    .public-visual-fallback i,
    .public-visual-fallback span {
      position: absolute;
      display: block;
      border-radius: 8px;
      background: rgba(255,255,255,.58);
      border: 1px solid rgba(20,20,20,.06);
    }
    .public-visual-fallback i:nth-child(1) { width: 72%; height: 12%; left: 13%; top: 22%; transform: rotate(-7deg); }
    .public-visual-fallback i:nth-child(2) { width: 48%; height: 16%; right: 7%; top: 48%; transform: rotate(8deg); }
    .public-visual-fallback i:nth-child(3) { width: 68%; height: 13%; left: 10%; bottom: 20%; transform: rotate(-3deg); }
    .public-visual-fallback span { left: 20%; bottom: 30%; width: 46%; height: 7px; background: var(--ink); opacity: .82; }
    .public-hero-card {
      position: absolute;
      right: 18px;
      bottom: 18px;
      width: min(360px, calc(100% - 36px));
      border-radius: 8px;
      background: rgba(255,255,255,.82);
      border: 1px solid rgba(255,255,255,.76);
      backdrop-filter: blur(16px);
      padding: 16px;
      box-shadow: 0 18px 50px rgba(20,20,20,.16);
    }
    .public-progress-head { display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; margin-bottom: 12px; }
    .public-progress-head span { color: var(--muted); font-size: 12px; }
    .public-progress-head strong { color: var(--red); font-size: 14px; text-align: right; }
    .public-progress-line { height: 10px; border-radius: 999px; background: #e5e1d9; overflow: hidden; margin-bottom: 13px; }
    .public-progress-line i { display: block; height: 100%; border-radius: inherit; background: linear-gradient(90deg, var(--red), var(--gold)); }
    .public-progress-meta { display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; }
    .public-progress-meta span { color: var(--muted); font-size: 11px; display: block; }
    .public-progress-meta b { display: block; margin-top: 4px; font-size: 16px; }
    .public-kpis {
      display: grid;
      grid-template-columns: repeat(6, minmax(120px, 1fr));
      gap: 12px;
      margin-bottom: 20px;
    }
    .public-kpi {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--white);
      padding: 16px;
      min-height: 118px;
    }
    .public-kpi span { color: var(--muted); font-size: 12px; display: block; }
    .public-kpi strong { font-size: 28px; line-height: 1; display: block; margin: 11px 0 7px; }
    .public-kpi small { color: var(--muted); line-height: 1.35; display: block; }
    .public-kpi.hot { border-top: 4px solid var(--red); }
    .public-kpi.gold { border-top: 4px solid var(--gold); }
    .public-grid {
      display: grid;
      grid-template-columns: minmax(320px, .94fr) minmax(420px, 1.06fr);
      gap: 20px;
      margin-bottom: 20px;
      align-items: start;
    }
    .public-panel {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--white);
      box-shadow: 0 14px 44px rgba(20,20,20,.08);
      padding: 20px;
      min-width: 0;
    }
    .public-panel h2 { margin: 0 0 14px; font-size: 16px; letter-spacing: 0; }
    .public-panel-head {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 14px;
    }
    .public-panel-head h2 { margin: 0; }
    .public-panel-head span { color: var(--muted); font-size: 12px; white-space: nowrap; }
    .public-activity { list-style: none; padding: 0; margin: 0; display: grid; gap: 11px; }
    .public-activity li { display: grid; grid-template-columns: 42px 1fr; gap: 12px; align-items: start; }
    .public-activity b { width: 38px; height: 38px; display: grid; place-items: center; border-radius: 8px; background: var(--ink); color: #fff; font-size: 12px; }
    .public-activity span { color: var(--red); font-size: 12px; font-weight: 850; }
    .public-activity p { margin: 3px 0 0; line-height: 1.45; color: var(--ink); }
    .public-finance { display: grid; gap: 10px; }
    .public-finance-row {
      display: grid;
      grid-template-columns: minmax(95px, 1fr) minmax(160px, 2fr);
      gap: 14px;
      align-items: center;
      padding: 10px 0;
      border-bottom: 1px solid var(--line);
    }
    .public-finance-row:last-child { border-bottom: 0; }
    .public-finance-row span { color: var(--muted); font-size: 12px; }
    .public-finance-row strong { display: block; margin-top: 2px; }
    .public-bar { height: 9px; border-radius: 999px; background: #ebe7de; overflow: hidden; }
    .public-bar i { display: block; height: 100%; border-radius: inherit; background: var(--gold); }
    .public-photo-grid {
      display: grid;
      grid-template-columns: 1.3fr repeat(2, .85fr);
      gap: 12px;
      margin-top: 8px;
    }
    .public-photo-grid figure { margin: 0; min-width: 0; }
    .public-photo-grid img, .public-photo-grid figure > div {
      width: 100%;
      aspect-ratio: 1.12;
      object-fit: cover;
      display: block;
      border-radius: 8px;
      background: linear-gradient(135deg, #e8e8e4, #cfcfca);
      border: 1px solid var(--line);
    }
    .public-photo-grid figure:first-child { grid-row: span 2; }
    .public-photo-grid figure:first-child img,
    .public-photo-grid figure:first-child > div { height: 100%; aspect-ratio: auto; }
    .public-photo-grid figcaption { color: var(--muted); font-size: 12px; margin-top: 7px; line-height: 1.35; }
    .public-problem {
      display: grid;
      grid-template-columns: 84px 1fr;
      gap: 16px;
      align-items: center;
    }
    .public-problem strong {
      width: 84px;
      aspect-ratio: 1;
      display: grid;
      place-items: center;
      border-radius: 8px;
      background: #f7ebe5;
      color: var(--red);
      font-size: 24px;
      letter-spacing: 0;
    }
    .public-problem span { color: var(--red); font-size: 12px; text-transform: uppercase; font-weight: 850; }
    .public-problem p { margin: 5px 0 0; line-height: 1.48; color: var(--ink); }
    .public-empty { color: var(--muted); margin: 0; line-height: 1.5; }
    @media (max-width: 1160px) {
      .public-hero, .public-grid { grid-template-columns: 1fr; }
      .public-exec-band { grid-template-columns: 1fr 1fr; }
      .public-kpis { grid-template-columns: repeat(3, 1fr); }
      .public-visual { min-height: 440px; }
    }
    @media (max-width: 720px) {
      .public-nav { align-items: flex-start; flex-direction: column; }
      .public-links { order: 3; }
      main { padding: 18px 14px 42px; }
      .public-hero { min-height: 0; }
      .public-hero-copy { padding: 22px; }
      .public-hero-bottom, .public-kpis, .public-progress-meta, .public-exec-band { grid-template-columns: 1fr 1fr; }
      .public-visual { min-height: 360px; }
      .public-photo-grid { grid-template-columns: 1fr 1fr; }
      .public-photo-grid figure:first-child { grid-column: 1 / -1; min-height: 260px; }
      .public-finance-row { grid-template-columns: 1fr; gap: 8px; }
    }
    @media (max-width: 460px) {
      .public-kpis, .public-hero-bottom, .public-progress-meta, .public-exec-band { grid-template-columns: 1fr; }
      .public-hero h1 { font-size: 28px; }
      .public-problem { grid-template-columns: 1fr; }
      .public-problem strong { width: 66px; }
    }
    """


def _render_public_activity(items: list[dict]) -> str:
    if not items:
        return '<p class="public-empty">No field activity has been posted for the active report date yet.</p>'
    html = []
    for item in items[:5]:
        html.append(
            f"""
            <li>
              <b>{escape(str(item["seq"]).zfill(2))}</b>
              <div>
                <span>{escape(item["type"])}</span>
                <p>{escape(item["text"])}</p>
              </div>
            </li>
            """
        )
    return '<ol class="public-activity">' + "".join(html) + "</ol>"


def _render_public_photos(photos: list[dict]) -> str:
    cards = []
    for photo in photos[:5]:
        visual = (
            f'<img src="{escape(str(photo["url"]))}" alt="{escape(photo["caption"])}">'
            if photo.get("url")
            else "<div></div>"
        )
        cards.append(
            f"""
            <figure>
              {visual}
              <figcaption>{escape(photo["caption"])}</figcaption>
            </figure>
            """
        )
    while len(cards) < 5:
        cards.append("<figure><div></div><figcaption>Field photo pending</figcaption></figure>")
    return '<div class="public-photo-grid">' + "".join(cards) + "</div>"


def _render_public_finance(finance: dict) -> str:
    rows = [
        ("Budget", finance.get("budget_label", "-"), 100),
        ("Disbursed", finance.get("spent_label", "-"), finance.get("percent", 0)),
        ("Balance", finance.get("remaining_label", "-"), max(0, 100 - float(finance.get("percent") or 0))),
    ]
    html = []
    for label, value, width in rows:
        html.append(
            f"""
            <div class="public-finance-row">
              <div><span>{escape(label)}</span><strong>{escape(str(value))}</strong></div>
              <div class="public-bar"><i style="width:{max(0, min(100, float(width))):.1f}%"></i></div>
            </div>
            """
        )
    return '<div class="public-finance">' + "".join(html) + "</div>"


def _public_api_payload(data: dict) -> dict:
    return {
        "generated_at": data["generated_at"],
        "project_name": data["env"]["project_name"],
        "field_context": data["field_context"],
        "today_activities": data["today_activities"],
        "financial_report": data["financial_report"],
        "progress_report": data["progress_report"],
        "problem_report": data["problem_report"],
        "recent_photos": data["recent_photos"],
    }


def _render_public_page(data: dict) -> str:
    ctx = data["field_context"]
    finance = data["financial_report"]
    progress = data["progress_report"]
    problem = data["problem_report"]
    photos = data["recent_photos"]
    hero_photo = next((p for p in photos if p.get("url")), None)
    visual = (
        f'<img src="{escape(str(hero_photo["url"]))}" alt="{escape(hero_photo["caption"])}">'
        if hero_photo
        else '<div class="public-visual-fallback"><i></i><i></i><i></i><span></span></div>'
    )
    status_note = "Ahead of plan" if progress["variance_percent"] > 0 else "Behind plan" if progress["variance_percent"] < 0 else "On plan"
    public_watch_status = "Clear" if int(problem.get("count") or 0) == 0 else "Under Review"
    public_watch_mark = "OK" if int(problem.get("count") or 0) == 0 else "TRACK"
    public_watch_text = (
        "No public-facing risk signal is currently shown on this page."
        if int(problem.get("count") or 0) == 0
        else "Field actions are being tracked internally by the project control team."
    )
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Hua Ro Project Status</title>
  <style>{_public_css()}</style>
</head>
<body>
<div class="public-shell">
  <header class="public-nav">
    <div class="public-brand">
      <div class="public-mark" aria-hidden="true"></div>
      <div>
        <b>Hua Ro Project Status</b>
        <span>Public construction dashboard</span>
      </div>
    </div>
    <nav class="public-links" aria-label="Public sections">
      <a href="#progress">Progress</a>
      <a href="#finance">Finance</a>
      <a href="#field">Field</a>
      <a href="#photos">Photos</a>
    </nav>
    <div class="public-pills">
      <span class="public-pill"><i class="public-dot"></i> Live report</span>
      <span class="public-pill">Updated {escape(data["generated_at"])}</span>
    </div>
  </header>

  <main>
    <section class="public-hero">
      <div class="public-hero-copy">
        <div>
          <span class="public-eyebrow">Project Overview</span>
          <h1>{escape(PROJECT_DISPLAY_NAME)}</h1>
          <p class="public-lead">{escape(ctx["latest_text"])}</p>
          <div class="public-source">
            <span>Source: Construction Plan workbook</span>
            <span>Daily field report</span>
            <span>LINE photo evidence</span>
          </div>
        </div>
        <div class="public-hero-bottom">
          <div class="public-signal">
            <span>Current field date</span>
            <strong>{escape(ctx["work_date"] or "Waiting for report")}</strong>
          </div>
          <div class="public-signal">
            <span>Site condition</span>
            <strong>{escape(ctx["weather"])} · Water {escape(ctx["water"] or "-")}</strong>
          </div>
        </div>
      </div>
      <div class="public-visual">
        {visual}
        <div class="public-hero-card">
          <div class="public-progress-head">
            <span>Construction progress</span>
            <strong>{escape(status_note)}</strong>
          </div>
          <div class="public-progress-line"><i style="width:{max(0, min(100, int(progress["percent"])))}%"></i></div>
          <div class="public-progress-meta">
            <div><span>Plan</span><b>{escape(_percent_label(progress["plan_percent"]))}</b></div>
            <div><span>Actual</span><b>{escape(_percent_label(progress["actual_percent"]))}</b></div>
            <div><span>Variance</span><b>{escape(_signed_percent_label(progress["variance_percent"]))}</b></div>
          </div>
        </div>
      </div>
    </section>

    <section class="public-exec-band" aria-label="Executive summary">
      <article>
        <span>Executive Read</span>
        <strong>{escape(status_note)} / {escape(_signed_percent_label(progress["variance_percent"]))}</strong>
        <p>{escape(progress["period_label"])} from the latest approved progress table.</p>
      </article>
      <article><span>Contract Value</span><strong>{escape(finance["budget_label"])}</strong></article>
      <article><span>Paid To Date</span><strong>{escape(finance["spent_label"])}</strong></article>
      <article><span>Field Force</span><strong>{escape(str(ctx["workers"]))} workers</strong></article>
    </section>

    <section id="progress" class="public-kpis" aria-label="Project indicators">
      <article class="public-kpi hot"><span>Actual Progress</span><strong>{escape(_percent_label(progress["actual_percent"]))}</strong><small>{escape(progress["period_label"])}</small></article>
      <article class="public-kpi"><span>Plan</span><strong>{escape(_percent_label(progress["plan_percent"]))}</strong><small>Week {escape(str(progress["week_no"] or "-"))}</small></article>
      <article class="public-kpi gold"><span>Disbursed</span><strong>{escape(str(finance["headline"]).split()[0])}</strong><small>{escape(finance["spent_label"])} paid</small></article>
      <article class="public-kpi"><span>Workers</span><strong>{escape(str(ctx["workers"]))}</strong><small>Latest daily report</small></article>
      <article class="public-kpi"><span>Engineers</span><strong>{escape(str(ctx["engineers"]))}</strong><small>Latest daily report</small></article>
      <article class="public-kpi gold"><span>Machines</span><strong>{escape(str(ctx["equipment_count"]))}</strong><small>Latest daily report</small></article>
    </section>

    <section id="field" class="public-grid">
      <div class="public-panel">
        <div class="public-panel-head"><h2>Today On Site</h2><span>{escape(ctx["work_date"] or "Live")}</span></div>
        {_render_public_activity(data["today_activities"])}
      </div>
      <div id="finance" class="public-panel">
        <div class="public-panel-head"><h2>Financial Snapshot</h2><span>{escape(finance["note"])}</span></div>
        {_render_public_finance(finance)}
      </div>
    </section>

    <section class="public-grid">
      <div class="public-panel">
        <div class="public-panel-head"><h2>Governance Watch</h2><span>{escape(public_watch_status)}</span></div>
        <div class="public-problem">
          <strong>{escape(public_watch_mark)}</strong>
          <div>
            <span>{escape(public_watch_status)}</span>
            <p>{escape(public_watch_text)}</p>
          </div>
        </div>
      </div>
      <div id="photos" class="public-panel">
        <div class="public-panel-head"><h2>Recent Field Photos</h2><span>Latest evidence</span></div>
        {_render_public_photos(photos)}
      </div>
    </section>
  </main>
</div>
</body>
</html>"""
    return html


def _render_admin_page(data: dict, token: str) -> str:
    token_q = quote(token, safe="")
    env = data["env"]
    metrics = data["metrics"]
    db_dot = "ok" if data["db_ok"] else "bad"
    weekly_dot = "ok" if env["weekly_enabled"] else "warn"
    query_errors = data.get("query_errors") or []
    error_html = ""
    if data.get("db_error"):
        error_html += f'<div class="error">{escape(data["db_error"])}</div>'
    if query_errors:
        joined = " | ".join(_short(err, 180) for err in query_errors[:3])
        error_html += f'<div class="error">{escape(joined)}</div>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Hua Ro Project Command Dashboard</title>
  <style>{_page_css()}</style>
</head>
<body>
<div class="shell">
  <header class="topbar">
    <div class="brand">
      <div class="brand-mark" aria-hidden="true"></div>
      <div>
        <h1>Project Command Dashboard</h1>
        <span>LINE bot, reports, Supabase and Railway operations</span>
      </div>
    </div>
    <div class="top-actions">
      <span class="pill"><i class="dot {db_dot}"></i> Supabase</span>
      <span class="pill"><i class="dot {weekly_dot}"></i> Weekly {escape(env["weekly_format"])}</span>
      <span class="pill">Updated {escape(data["generated_at"])}</span>
    </div>
  </header>

  <main>
    {_render_admin_command_brief(data)}
    {_render_site_visual(data)}
    {_render_cockpit_cards(data)}
    {_render_problem_board(data, token_q)}

    <section class="metrics" aria-label="Key metrics">
      {_render_metric("Today Messages", metrics["today_messages"], "LINE entries dated today", "#b91c1c")}
      {_render_metric("7 Day Messages", metrics["week_messages"], "Recent report traffic", "#f59e0b")}
      {_render_metric("Active Users", metrics["active_users"], "Unique LINE user ids", "#7f1d1d")}
      {_render_metric("Photos", metrics["photos_30d"], "Stored progress images", "#dc2626")}
      {_render_metric("Workers 7D", metrics["worker_total_7d"], "Sum from daily reports", "#facc15")}
      {_render_metric("Open Issues", metrics["open_issues"], "Problem board items", "#18110f")}
    </section>

    <section class="dashboard-grid">
      <div class="section">
        <h2>Today's Activity</h2>
        {_render_today_activity(data["today_activities"])}
      </div>
      <div class="chart-panel">
        <div class="split-head">
          <h2>Fourteen Day Field Signal</h2>
          <span class="pill">Messages + workforce</span>
        </div>
        {_render_report_chart(data["daily_series"])}
      </div>
    </section>

    <section class="grid-2">
      <div class="section">
        <h2>Recent Field Photos</h2>
        {_render_photo_strip(data["recent_photos"])}
      </div>
      <div class="section">
        <h2>Water Level Trend</h2>
        {_render_waterline(data["water_points"])}
      </div>
    </section>

    <section class="section" style="margin-top:18px">
      <div class="split-head">
        <h2>Operational Topology</h2>
        <form action="/admin/trigger/weekly?token={token_q}" method="post">
          <button class="primary" type="submit">Run Weekly Report</button>
        </form>
      </div>
      <div class="flow">{_render_flow(data)}</div>
      {error_html}
    </section>

    <section class="grid-2" style="margin-top:18px">
      <div class="section">
        <h2>Activity Mix</h2>
        {_render_activity_chart(data["activity_top"])}
      </div>
      <div class="section">
        <h2>Latest Daily Snapshot</h2>
        {_render_latest_daily(data["latest_daily"])}
      </div>
    </section>

    <section class="section" style="margin-top:18px">
      <h2>Recent LINE Activity</h2>
      {_render_recent(data["recent_events"])}
    </section>

    <section class="section" style="margin-top:18px">
      <h2>Data Workbooks</h2>
      <table>
        <thead>
          <tr><th>File</th><th>Status</th><th>Size</th><th>Modified</th><th>Controls</th></tr>
        </thead>
        <tbody>{_render_files(data["files"], token_q)}</tbody>
      </table>
      <p class="footnote">The admin token stays server side. Supabase and LINE credentials are never printed on this page.</p>
    </section>
  </main>
</div>
</body>
</html>"""
    return html


def _admin_redirect(token: str) -> RedirectResponse:
    return RedirectResponse(url=f"/admin?token={quote(token, safe='')}", status_code=303)


def _clean_form_text(value: str | None) -> str | None:
    text = " ".join(str(value or "").split())
    return text or None


def _clean_form_date(value: str | None) -> str | None:
    parsed = _parse_date(value)
    return parsed.isoformat() if parsed else None


def _clean_status(value: str | None) -> str:
    value = str(value or "open").strip().lower()
    return value if value in {"open", "in_progress", "waiting", "resolved", "closed"} else "open"


def _clean_impact(value: str | None) -> str:
    value = str(value or "medium").strip().lower()
    return value if value in {"low", "medium", "high", "critical"} else "medium"


@public_router.get("/project", response_class=HTMLResponse)
async def public_project_home():
    return HTMLResponse(_render_public_page(_collect_dashboard()))


@public_router.get("/project/api")
async def public_project_api():
    return JSONResponse(_public_api_payload(_collect_dashboard()))


@router.post("/issues/create")
async def admin_issue_create(
    token: str = "",
    title: str = Form(...),
    area: str = Form(""),
    owner: str = Form(""),
    due_date: str = Form(""),
    impact: str = Form("medium"),
    next_action: str = Form(""),
):
    _check_token(token)
    client, err = _get_supabase_client()
    if client is None:
        raise HTTPException(503, err or "Supabase is not configured.")
    payload = {
        "title": _clean_form_text(title),
        "area": _clean_form_text(area),
        "owner": _clean_form_text(owner),
        "due_date": _clean_form_date(due_date),
        "impact": _clean_impact(impact),
        "next_action": _clean_form_text(next_action),
        "status": "open",
        "source_channel": "admin",
        "updated_at": _now_bkk().astimezone(timezone.utc).isoformat(),
    }
    try:
        client.table("project_issues").insert(payload).execute()
    except Exception as exc:
        raise HTTPException(500, f"Cannot create issue: {exc}")
    return _admin_redirect(token)


@router.post("/issues/{issue_id}/update")
async def admin_issue_update(
    issue_id: int,
    token: str = "",
    owner: str = Form(""),
    due_date: str = Form(""),
    status: str = Form("open"),
    next_action: str = Form(""),
):
    _check_token(token)
    client, err = _get_supabase_client()
    if client is None:
        raise HTTPException(503, err or "Supabase is not configured.")
    payload = {
        "owner": _clean_form_text(owner),
        "due_date": _clean_form_date(due_date),
        "status": _clean_status(status),
        "next_action": _clean_form_text(next_action),
        "updated_at": _now_bkk().astimezone(timezone.utc).isoformat(),
    }
    try:
        client.table("project_issues").update(payload).eq("id", issue_id).execute()
    except Exception as exc:
        raise HTTPException(500, f"Cannot update issue: {exc}")
    return _admin_redirect(token)


@router.post("/issues/{issue_id}/comment")
async def admin_issue_comment(
    issue_id: int,
    token: str = "",
    comment: str = Form(...),
):
    _check_token(token)
    client, err = _get_supabase_client()
    if client is None:
        raise HTTPException(503, err or "Supabase is not configured.")
    text = _clean_form_text(comment)
    if text:
        try:
            client.table("project_issue_comments").insert(
                {"issue_id": issue_id, "author": "Admin", "comment": text}
            ).execute()
            client.table("project_issues").update(
                {"updated_at": _now_bkk().astimezone(timezone.utc).isoformat()}
            ).eq("id", issue_id).execute()
        except Exception as exc:
            raise HTTPException(500, f"Cannot save comment: {exc}")
    return _admin_redirect(token)


@router.get("", response_class=HTMLResponse)
async def admin_home(token: str = ""):
    _check_token(token)
    return HTMLResponse(_render_admin_page(_collect_dashboard(), token))


@router.get("/api/overview")
async def admin_api_overview(token: str = ""):
    _check_token(token)
    data = _collect_dashboard()
    public = {
        "generated_at": data["generated_at"],
        "db_ok": data["db_ok"],
        "query_errors": data["query_errors"],
        "health_score": data["health_score"],
        "metrics": data["metrics"],
        "daily_series": data["daily_series"],
        "activity_top": data["activity_top"],
        "water_points": data["water_points"],
        "files": data["files"],
        "field_context": data["field_context"],
        "today_activities": data["today_activities"],
        "financial_report": data["financial_report"],
        "progress_report": data["progress_report"],
        "problem_report": data["problem_report"],
        "issue_board": {
            "open_count": data["issue_board"]["open_count"],
            "total_count": data["issue_board"]["total_count"],
            "overdue_count": data["issue_board"]["overdue_count"],
            "due_soon_count": data["issue_board"]["due_soon_count"],
            "by_status": data["issue_board"]["by_status"],
            "by_impact": data["issue_board"]["by_impact"],
            "open_issues": data["issue_board"]["open_issues"],
        },
        "recent_photos": data["recent_photos"],
    }
    return JSONResponse(public)


@router.get("/download/{kind}")
async def admin_download(kind: str, token: str = ""):
    _check_token(token)
    if kind not in DATA_FILES:
        raise HTTPException(404, f"Unknown data file kind: {kind}")
    path, _ = DATA_FILES[kind]
    if not path.exists():
        raise HTTPException(404, f"Missing file: {path.name}")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/upload/{kind}")
async def admin_upload(kind: str, token: str = "", file: UploadFile = File(...)):
    _check_token(token)
    if kind not in DATA_FILES:
        raise HTTPException(404, f"Unknown data file kind: {kind}")
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx uploads are supported.")

    path, _ = DATA_FILES[kind]
    content = await file.read()
    if path.exists():
        backup = path.with_suffix(f".{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak.xlsx")
        path.rename(backup)
    path.write_bytes(content)
    return RedirectResponse(f"/admin?token={quote(token, safe='')}", status_code=303)


@router.post("/trigger/weekly")
async def admin_trigger_weekly(token: str = ""):
    _check_token(token)
    try:
        from scheduler import run_weekly_for_current_week

        await run_weekly_for_current_week()
        return JSONResponse({"status": "ok", "message": "Weekly report triggered. Check LINE."})
    except Exception as exc:
        return JSONResponse({"status": "error", "error": str(exc)}, status_code=500)


@router.get("/check_fonts")
async def admin_check_fonts(token: str = ""):
    _check_token(token)
    result = {"thai_fonts": [], "soffice": "not_found", "errors": []}
    try:
        probe = subprocess.run(["fc-list", ":lang=th"], capture_output=True, timeout=10)
        if probe.returncode == 0:
            lines = probe.stdout.decode("utf-8", errors="ignore").splitlines()
            result["thai_fonts"] = sorted(set(line.split(":")[1].strip() for line in lines if ":" in line))
        else:
            result["errors"].append(f"fc-list failed: {probe.stderr.decode(errors='ignore')[:200]}")
    except Exception as exc:
        result["errors"].append(f"fc-list error: {exc}")

    for command in ("soffice", "libreoffice", "/usr/bin/soffice"):
        try:
            probe = subprocess.run([command, "--version"], capture_output=True, timeout=10)
            if probe.returncode == 0:
                result["soffice"] = probe.stdout.decode("utf-8", errors="ignore").strip()
                break
        except Exception:
            continue
    return JSONResponse(result)
